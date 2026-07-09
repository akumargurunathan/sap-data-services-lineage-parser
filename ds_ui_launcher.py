#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DS XML Engine UI Launcher

Tkinter-based GUI for SAP Data Services XML Lineage Extraction
- Select input directory/file containing DS XML files
- Select output directory for results
- Run extraction with real-time log display
- View results and lineage mappings
"""

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import os
import re
import pandas as pd
import sys
import threading
import queue
import logging
import json
import time

# Direct imports from ds_engine to ensure alignment with engine functions
from ds_engine.table_lineage_extractor import extract_table_lineage

# ---------------------------------------------------------------------------
# Self-contained interactive HTML lineage graph — supports 1..N target tabs.
# Placeholder replaced at runtime: %%DATASETS%%  (JSON array of {tgt, recs})
# ---------------------------------------------------------------------------
_LINEAGE_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Lineage</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:#f8f8f6;color:#1a1a18;padding:16px}
@media(prefers-color-scheme:dark){body{background:#1a1a18;color:#e0ddd4}}
h1{font-size:14px;font-weight:500;margin-bottom:8px}
@media(prefers-color-scheme:dark){h1{color:#e0ddd4}}
.legend{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:8px;align-items:center;font-size:11px;color:#5f5e5a}
.lchip{display:flex;align-items:center;gap:5px}
.ldot{width:10px;height:10px;border-radius:2px;flex-shrink:0}
.canvas-wrap{width:100%;overflow-x:auto;overflow-y:auto;border:1px solid #d3d1c7;border-radius:0 8px 8px 8px;background:#fff;max-height:68vh}
@media(prefers-color-scheme:dark){.canvas-wrap{border-color:#444441;background:#2c2c2a}}
canvas{display:block}
.detail{margin-top:10px;padding:10px 14px;border-radius:8px;background:#f1efe8;border:1px solid #d3d1c7;font-size:12px;min-height:60px;color:#5f5e5a;line-height:1.8}
@media(prefers-color-scheme:dark){.detail{background:#2c2c2a;border-color:#444441;color:#888780}}
.detail b{color:#1a1a18;font-weight:500}
@media(prefers-color-scheme:dark){.detail b{color:#e0ddd4}}
.tag{display:inline-block;font-size:10px;padding:1px 6px;border-radius:3px;margin:1px 2px;font-family:monospace}
.tabs{display:flex;gap:0;flex-wrap:wrap;margin-bottom:-1px;position:relative;z-index:1}
.tabs.hidden{display:none}
.tab-btn{padding:5px 14px;border-radius:6px 6px 0 0;border:1px solid #d3d1c7;border-bottom:none;font-size:12px;cursor:pointer;background:#f1efe8;color:#5f5e5a;white-space:nowrap;font-weight:400}
.tab-btn:hover{background:#e8e6df}
.tab-btn.active{background:#fff;color:#1a1a18;font-weight:600;border-bottom:1px solid #fff}
@media(prefers-color-scheme:dark){.tab-btn{background:#2c2c2a;border-color:#444441;color:#888780}.tab-btn:hover{background:#383836}.tab-btn.active{background:#1a1a18;color:#e0ddd4;border-bottom:1px solid #1a1a18}}
.toolbar{display:flex;gap:8px;align-items:center;margin-bottom:6px;flex-wrap:wrap;border:1px solid #d3d1c7;border-radius:0 8px 8px 8px;padding:6px 10px;background:#fff}
@media(prefers-color-scheme:dark){.toolbar{background:#1a1a18;border-color:#444441}}
.tabs.hidden~.canvas-wrap,.tabs.hidden~.toolbar{border-radius:8px}
#search{flex:1;min-width:160px;max-width:300px;padding:4px 10px;border:1px solid #d3d1c7;border-radius:6px;font-size:12px;background:#fff;color:#1a1a18;outline:none}
#search:focus{border-color:#378ADD;box-shadow:0 0 0 2px #378ADD22}
@media(prefers-color-scheme:dark){#search{background:#2c2c2a;border-color:#444441;color:#e0ddd4}}
.btn-sm{padding:3px 10px;border-radius:6px;border:1px solid #d3d1c7;font-size:11px;cursor:pointer;background:#f1efe8;color:#5f5e5a;white-space:nowrap}
.btn-sm:hover{background:#e8e6df}
@media(prefers-color-scheme:dark){.btn-sm{background:#2c2c2a;border-color:#444441;color:#888780}.btn-sm:hover{background:#383836}}
.btn-export{border-color:#1D9E7588;color:#0F6E56;background:#E1F5EE}
.btn-export:hover{background:#C8EFE2}
@media(prefers-color-scheme:dark){.btn-export{background:#085041;border-color:#1D9E75;color:#9FE1CB}}
.spot-badge{display:inline-flex;align-items:center;gap:6px;padding:3px 10px;border-radius:12px;background:#378ADD18;border:1px solid #378ADD55;color:#378ADD;font-size:11px;font-weight:500}
.spot-badge button{background:none;border:none;cursor:pointer;color:#378ADD;font-size:14px;line-height:1;padding:0}
.ctx-menu{position:fixed;background:#fff;border:1px solid #d3d1c7;border-radius:8px;box-shadow:0 4px 20px rgba(0,0,0,0.14);padding:4px 0;min-width:170px;z-index:9999;font-size:12px}
@media(prefers-color-scheme:dark){.ctx-menu{background:#2c2c2a;border-color:#444441}}
.ctx-hdr{padding:6px 14px 5px;font-size:10px;color:#888780;border-bottom:1px solid #d3d1c7;margin-bottom:3px}
@media(prefers-color-scheme:dark){.ctx-hdr{border-color:#444441}}
.ctx-item{padding:7px 14px;cursor:pointer;color:#1a1a18;display:flex;align-items:center;gap:8px}
.ctx-item:hover{background:#f1efe8}
@media(prefers-color-scheme:dark){.ctx-item{color:#e0ddd4}.ctx-item:hover{background:#383836}}
.ctx-sep{height:1px;background:#d3d1c7;margin:3px 0}
@media(prefers-color-scheme:dark){.ctx-sep{background:#444441}}
/* ── Column detail panel ── */
.col-panel{position:fixed;top:0;right:-430px;width:430px;height:100vh;background:#fff;border-left:1px solid #d3d1c7;box-shadow:-4px 0 24px rgba(0,0,0,0.13);z-index:10000;display:flex;flex-direction:column;transition:right .22s ease}
.col-panel.open{right:0}
@media(prefers-color-scheme:dark){.col-panel{background:#1e1e1c;border-color:#444441}}
.sql-panel{position:fixed;top:0;right:-520px;width:520px;height:100vh;background:#fff;border-left:2px solid #D84315;box-shadow:-6px 0 28px rgba(216,67,21,0.15);z-index:10001;display:flex;flex-direction:column;transition:right .22s ease}
.sql-panel.open{right:0}
@media(prefers-color-scheme:dark){.sql-panel{background:#1e1e1c;border-color:#BF360C}}
.cp-hdr{padding:12px 14px 10px;border-bottom:1px solid #e8e6df;display:flex;align-items:flex-start;gap:8px;flex-shrink:0}
@media(prefers-color-scheme:dark){.cp-hdr{border-color:#444441}}
.cp-edge{flex:1;min-width:0}
.cp-edge-title{font-size:12px;font-weight:600;color:#1a1a18;word-break:break-all;display:flex;align-items:center;gap:4px;flex-wrap:wrap}
@media(prefers-color-scheme:dark){.cp-edge-title{color:#e0ddd4}}
.cp-edge-title .cp-arr{color:#378ADD;flex-shrink:0}
.cp-edge-sub{font-size:10px;color:#888780;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cp-close{background:none;border:none;cursor:pointer;color:#888780;font-size:20px;line-height:1;padding:0;flex-shrink:0;margin-top:-2px}
.cp-close:hover{color:#E53935}
.cp-tabs{display:flex;border-bottom:1px solid #d3d1c7;background:#f8f8f6;flex-shrink:0}
@media(prefers-color-scheme:dark){.cp-tabs{background:#2c2c2a;border-color:#444441}}
.cp-tab{flex:1;padding:8px 4px;border:none;background:none;cursor:pointer;font-size:11px;color:#888780;border-bottom:2px solid transparent;transition:color .15s,border-color .15s}
.cp-tab:hover{color:#1a1a18;background:#f1efe8}
.cp-tab.active{color:#378ADD;border-bottom-color:#378ADD;font-weight:600}
@media(prefers-color-scheme:dark){.cp-tab:hover{color:#e0ddd4;background:#383836}}
.cp-cnt{display:inline-block;min-width:16px;padding:0 4px;border-radius:8px;background:#e8e6df;font-size:9px;font-weight:700;margin-left:3px;vertical-align:middle}
@media(prefers-color-scheme:dark){.cp-cnt{background:#3a3a38;color:#c0bfb8}}
.cp-body{flex:1;overflow-y:auto}
.cp-empty{padding:28px 18px;text-align:center;color:#888780;font-size:12px;line-height:1.6}
.cp-df-hdr{padding:7px 14px 4px;font-size:9px;font-weight:700;color:#888780;text-transform:uppercase;letter-spacing:.5px;background:#f8f8f6;border-bottom:1px solid #e8e6df;position:sticky;top:0;z-index:1}
@media(prefers-color-scheme:dark){.cp-df-hdr{background:#2c2c2a;border-color:#444441}}
.cp-df-picker{padding:6px 12px 5px;border-bottom:1px solid #d3d1c7;background:#f8f8f6;display:flex;align-items:center;gap:8px;flex-shrink:0}
@media(prefers-color-scheme:dark){.cp-df-picker{background:#2c2c2a;border-color:#444441}}
.cp-df-picker-lbl{font-size:10px;font-weight:600;color:#888780;white-space:nowrap;text-transform:uppercase;letter-spacing:.4px}
.cp-df-sel{flex:1;font-size:11px;padding:3px 6px;border:1px solid #d3d1c7;border-radius:5px;background:#fff;color:#1a1a18;cursor:pointer;min-width:0}
@media(prefers-color-scheme:dark){.cp-df-sel{background:#1e1e1c;border-color:#555;color:#c0bfb8}}
.cp-search-wrap{padding:6px 10px 4px;position:sticky;top:0;background:#fff;border-bottom:1px solid #e8e6df;z-index:2}
@media(prefers-color-scheme:dark){.cp-search-wrap{background:#1e1e1c;border-color:#444441}}
.cp-search{width:100%;box-sizing:border-box;padding:4px 8px;border:1px solid #d3d1c7;border-radius:6px;font-size:11px;outline:none;background:#f8f8f6;color:#1a1a18}
.cp-search:focus{border-color:#6366f1;box-shadow:0 0 0 2px rgba(99,102,241,.15)}
@media(prefers-color-scheme:dark){.cp-search{background:#2c2c2a;border-color:#555;color:#c0bfb8}}
.cp-row{display:grid;grid-template-columns:22px 1fr 16px 1fr;gap:0 6px;align-items:center;padding:5px 12px;border-bottom:1px solid #f4f3ef;font-size:11px}
.cp-row.has-f{cursor:pointer}
.cp-row:hover{background:#f8f8f6}
@media(prefers-color-scheme:dark){.cp-row{border-color:#2a2a28}.cp-row:hover{background:#242422}}
.cp-badge{width:18px;height:18px;border-radius:3px;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:800;flex-shrink:0}
.cp-col{font-family:monospace;font-size:11px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#1a1a18}
@media(prefers-color-scheme:dark){.cp-col{color:#e0ddd4}}
.cp-col.dim{color:#aaa}
.cp-arr2{color:#378ADD;font-size:13px;text-align:center}
.cp-fx{grid-column:1/-1;padding:3px 8px 5px 28px;font-family:monospace;font-size:10px;color:#5f5e5a;background:#f8f7f3;border-radius:4px;margin:1px 2px 4px;word-break:break-all;line-height:1.5;display:none}
@media(prefers-color-scheme:dark){.cp-fx{background:#242422;color:#a0a098}}
.cp-fx.open{display:block}
.cp-cols-hdr{display:grid;grid-template-columns:22px 1fr 16px 1fr;gap:0 6px;padding:5px 12px 4px;font-size:9px;font-weight:700;letter-spacing:.5px;border-bottom:2px solid #c7f0df;background:#f0fcf6;color:#555}
@media(prefers-color-scheme:dark){.cp-cols-hdr{background:#0a2016;border-color:#1a5c35;color:#9fe1cb}}
.cp-col-src{background:rgba(29,158,117,.10);border-radius:3px;padding:1px 5px}
.cp-col-tgt{background:rgba(55,138,221,.10);border-radius:3px;padding:1px 5px}
.cp-fallback{padding:14px}
.cp-fallback p{font-size:12px;color:#888780;margin:0 0 10px;line-height:1.5}
.cp-fallback-sel{width:100%;padding:5px 8px;border:1px solid #d3d1c7;border-radius:5px;font-size:11px;background:#fff;color:#1a1a18;cursor:pointer}
@media(prefers-color-scheme:dark){.cp-fallback-sel{background:#1e1e1c;border-color:#555;color:#c0bfb8}}
.cp-join{padding:10px 14px;border-bottom:1px solid #f4f3ef;font-size:11px}
@media(prefers-color-scheme:dark){.cp-join{border-color:#2a2a28}}
.cp-jtype{display:inline-block;padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700;background:#E3F2FD;color:#1565C0;margin-bottom:5px}
@media(prefers-color-scheme:dark){.cp-jtype{background:#0D2944;color:#90CAF9}}
.cp-jdf{font-size:10px;color:#888780;margin-bottom:3px}
.cp-jcond{font-family:monospace;font-size:11px;color:#1a1a18;word-break:break-all}
@media(prefers-color-scheme:dark){.cp-jcond{color:#e0ddd4}}
.cp-flt{padding:10px 14px;border-bottom:1px solid #f4f3ef;font-size:11px}
@media(prefers-color-scheme:dark){.cp-flt{border-color:#2a2a28}}
.cp-flabel{font-size:9px;font-weight:700;color:#888780;text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px}
.cp-fcond{font-family:monospace;font-size:11px;color:#1a1a18;word-break:break-all}
@media(prefers-color-scheme:dark){.cp-fcond{color:#e0ddd4}}
/* Field trace / evolution panel */
.cp-trace-hdr{display:none;align-items:center;gap:8px;padding:6px 12px;background:#F5F0FF;border-bottom:1px solid #d3d1c7;font-size:11px;font-weight:600;color:#7C3AED;flex-shrink:0}
.cp-trace-hdr.open{display:flex}
@media(prefers-color-scheme:dark){.cp-trace-hdr{background:#1E0A3C;border-color:#444441;color:#B39DDB}}
.cp-trace-back{background:none;border:none;cursor:pointer;color:#7C3AED;font-size:18px;line-height:1;padding:0 2px;flex-shrink:0}
@media(prefers-color-scheme:dark){.cp-trace-back{color:#B39DDB}}
.cp-trace-lbl{flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:11px}
.cp-trace-badge{font-size:9px;background:#EDE7F6;color:#7C3AED;padding:1px 6px;border-radius:8px;border:1px solid #B39DDB55;white-space:nowrap;flex-shrink:0}
@media(prefers-color-scheme:dark){.cp-trace-badge{background:#2D1B69;color:#B39DDB;border-color:#7C3AED55}}
.cp-trace-btn{background:none;border:1px solid #7C3AED44;border-radius:3px;cursor:pointer;color:#7C3AED;font-size:9px;padding:0 4px;line-height:16px;margin-left:4px;flex-shrink:0}
.cp-trace-btn:hover{background:#EDE7F6}
@media(prefers-color-scheme:dark){.cp-trace-btn{color:#B39DDB;border-color:#7C3AED44}.cp-trace-btn:hover{background:#2D1B69}}
.cp-te-node{padding:6px 12px 4px;border-bottom:1px solid #f4f3ef}
@media(prefers-color-scheme:dark){.cp-te-node{border-color:#2a2a28}}
.cp-te-row{display:flex;align-items:center;gap:5px;font-size:11px;flex-wrap:wrap}
.cp-te-icon{font-size:12px;flex-shrink:0}
.cp-te-tbl{font-weight:600;color:#1a1a18;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
@media(prefers-color-scheme:dark){.cp-te-tbl{color:#e0ddd4}}
.cp-te-col{font-family:monospace;color:#7C3AED;font-size:11px}
@media(prefers-color-scheme:dark){.cp-te-col{color:#B39DDB}}
.cp-te-df{font-size:9px;color:#888780}
.cp-te-fx{font-family:monospace;font-size:10px;color:#5f5e5a;padding:3px 8px;background:#F5F0FF;border-radius:3px;word-break:break-all;margin-top:2px;line-height:1.5}
@media(prefers-color-scheme:dark){.cp-te-fx{background:#1E0A3C;color:#a0a098}}
.cp-te-terminal{font-style:italic;color:#1D9E75}
@media(prefers-color-scheme:dark){.cp-te-terminal{color:#9FE1CB}}
/* Field search */
#fieldSearch{flex:0 1 200px;min-width:130px;padding:4px 10px;border:1px solid #7C3AED44;border-radius:6px;font-size:12px;background:#fff;color:#1a1a18;outline:none}
#fieldSearch:focus{border-color:#7C3AED;box-shadow:0 0 0 2px #7C3AED22}
@media(prefers-color-scheme:dark){#fieldSearch{background:#2c2c2a;border-color:#7C3AED44;color:#e0ddd4}}
/* Grouped target-column rows in column panel */
.cp-tc-grp-hdr{display:grid;grid-template-columns:16px 1fr auto;gap:0 6px;padding:5px 12px;cursor:pointer;align-items:center;border-bottom:1px solid #f4f3ef;background:#f8f8f6}
.cp-tc-grp-hdr:hover{background:#f1efe8}
@media(prefers-color-scheme:dark){.cp-tc-grp-hdr{background:#252523;border-color:#2a2a28}.cp-tc-grp-hdr:hover{background:#2e2e2c}}
.cp-tc-grp-cnt{font-size:9px;background:#E3E8F0;color:#4A6080;padding:1px 6px;border-radius:8px;font-weight:700}
@media(prefers-color-scheme:dark){.cp-tc-grp-cnt{background:#1e2d40;color:#8aabcc}}
.cp-tc-grp-body{display:none}
.cp-tc-grp-body.open{display:block}
</style>
</head>
<body>
<h1>Data lineage &#8212; <span id="htbl"></span></h1>
<div class="legend">
  <span class="lchip"><span class="ldot" style="background:#1D9E75;border:1px solid #0F6E56"></span>DB Source</span>
  <span class="lchip"><span class="ldot" style="background:#795548;border:1px solid #4E342E"></span>SAP Source</span>
  <span class="lchip"><span class="ldot" style="background:#006064;border:1px solid #004D40"></span>Flat File</span>
  <span class="lchip"><span class="ldot" style="background:#7B1FA2;border:1px solid #4A0072"></span>Excel</span>
  <span class="lchip"><span class="ldot" style="background:#378ADD;border:1px solid #185FA5"></span>Intermediate</span>
  <span class="lchip"><span class="ldot" style="background:#BA7517;border:1px solid #854F0B"></span>Target</span>
  <span class="lchip"><span class="ldot" style="background:#D4537E;border:1px solid #993556"></span>Multi-job</span>
  <span class="lchip"><span class="ldot" style="background:#D84315;border:1px solid #BF360C"></span>SQL Transform</span>
  <span style="margin-left:auto;font-size:11px;color:#888780">Node click = path &middot; Edge click = columns &middot; Right-click = spotlight &middot; Dbl-click SQL = expand/collapse &middot; Esc = clear</span>
</div>
<div class="tabs" id="tabStrip"></div>
<div class="toolbar">
  <input type="text" id="search" placeholder="&#128269; Search table name..." autocomplete="off" spellcheck="false"/>
  <input type="text" id="fieldSearch" placeholder="&#128270; Search field&#8230;" autocomplete="off" spellcheck="false" title="Type a column/field name to highlight its lineage path (numbered edges, Ctrl+G)"/>
  <span id="spotBadge" style="display:none" class="spot-badge"><span id="spotLabel">Spotlight</span> <button onclick="clearSpotlight()" title="Clear spotlight (Esc)">&#x2715;</button></span>
  <span style="margin-left:auto;display:flex;gap:6px">
    <button class="btn-sm btn-export" onclick="exportCurrentTab()">&#11015; Export Tab</button>
    <button class="btn-sm btn-export" id="btnExpAll" onclick="exportAllTabs()">&#11015; Export All</button>
  </span>
</div>
<div class="canvas-wrap"><canvas id="c"></canvas></div>
<div class="detail" id="detail">Click any table node for path &amp; details &middot; Click an edge for column-level lineage.</div>
<div id="colPanel" class="col-panel">
  <div class="cp-hdr">
    <div class="cp-edge">
      <div class="cp-edge-title"><span id="cpSrc"></span><span class="cp-arr">&#8594;</span><span id="cpTgt"></span></div>
      <div class="cp-edge-sub" id="cpDf"></div>
    </div>
    <button class="cp-close" onclick="closeColPanel()" title="Close">&#x2715;</button>
  </div>
  <div id="cpDfPicker" class="cp-df-picker" style="display:none">
    <span class="cp-df-picker-lbl">Dataflow</span>
    <select id="cpDfSel" class="cp-df-sel" onchange="selectColDf(this.value)"></select>
  </div>
  <div class="cp-tabs" id="cpNormalTabs">
    <button class="cp-tab active" id="cpTabCols" onclick="showCpTab('cols')">Columns <span class="cp-cnt" id="cpCntCols">0</span></button>
    <button class="cp-tab" id="cpTabJoins" onclick="showCpTab('joins')">Joins <span class="cp-cnt" id="cpCntJoins">0</span></button>
    <button class="cp-tab" id="cpTabFilters" onclick="showCpTab('filters')">Filters <span class="cp-cnt" id="cpCntFilters">0</span></button>
  </div>
  <div class="cp-trace-hdr" id="cpTraceHdr">
    <button class="cp-trace-back" onclick="exitTraceMode()" title="Back to column mappings">&#8592;</button>
    <span class="cp-trace-lbl" id="cpTraceLbl">Field evolution</span>
    <span class="cp-trace-badge">&#9672; canvas traced</span>
  </div>
  <div id="cpBody" class="cp-body"></div>
</div>

<div id="sqlPanel" class="sql-panel">
  <div class="cp-hdr" style="background:#FBE9E7;border-bottom:1px solid #FFCCBC">
    <div class="cp-edge" style="flex:1;min-width:0">
      <div class="cp-edge-title"><span style="color:#D84315;font-weight:700">&#9660; SQL Transform</span>&nbsp;&nbsp;<b id="sqlPanelName" style="color:#3E2723;font-family:Consolas,monospace"></b></div>
      <div class="cp-edge-sub" id="sqlPanelCtx" style="color:#795548"></div>
    </div>
    <button class="cp-close" onclick="closeSqlPanel()" title="Close" style="color:#D84315">&#x2715;</button>
  </div>
  <div id="sqlSrcSection" style="padding:10px 14px 8px;border-bottom:1px solid #FFCCBC;flex-shrink:0;background:#FFF3E0"></div>
  <div id="sqlBody" style="flex:1;overflow-y:auto;padding:14px 16px"></div>
</div>

<script>
const DATASETS=%%DATASETS%%;

let RECS,TGT,MAX_HOP,EDGES,NODE_TYPE,LAYERS,positions,layerXs,totalW,totalH;
let nodeJobs,nodeDFs,nodeWFs,multiJob;
let _outEdges={},_inEdges={},edgeSlots=[];
let currentTab=0;
let hov=null,sel=null;
let pathNodes=new Set(),pathEdgeKeys=new Set();
let upNodes=new Set(),upEdgeKeys=new Set();
let dnNodes=new Set(),dnEdgeKeys=new Set();
let spotDir=null,searchStr='',_ctxTarget=null;
let SQL_MEMBERS={},collapsedSqlNodes=new Set(),expandedSqlNodes=new Set();
let fieldSearchStr='',fieldSearchNodes=new Set(),fieldSearchEdges=new Map();

const NW=180,NH=32,HGAP=60,VGAP=7,PAD=16,LHDR=20;
const canvas=document.getElementById('c');
const ctx=canvas.getContext('2d');
const isDark=matchMedia('(prefers-color-scheme:dark)').matches;
const C={
  src:  isDark?{f:'#085041',s:'#1D9E75',t:'#9FE1CB'}:{f:'#E1F5EE',s:'#1D9E75',t:'#085041'},
  sap:  isDark?{f:'#3E2723',s:'#795548',t:'#FFE082'}:{f:'#EFEBE9',s:'#4E342E',t:'#3E2723'},
  file: isDark?{f:'#004D40',s:'#00695C',t:'#B2EBF2'}:{f:'#E0F2F1',s:'#006064',t:'#004D40'},
  excel:isDark?{f:'#2D0A42',s:'#AB47BC',t:'#E1BEE7'}:{f:'#F3E5F5',s:'#7B1FA2',t:'#2D0A42'},
  xfm:  isDark?{f:'#042C53',s:'#378ADD',t:'#B5D4F4'}:{f:'#E6F1FB',s:'#378ADD',t:'#042C53'},
  tgt:  isDark?{f:'#412402',s:'#BA7517',t:'#FAC775'}:{f:'#FAEEDA',s:'#BA7517',t:'#412402'},
  multi:isDark?{f:'#4B1528',s:'#D4537E',t:'#F4C0D1'}:{f:'#FBEAF0',s:'#D4537E',t:'#72243E'},
  sql:  isDark?{f:'#3E0A00',s:'#F4511E',t:'#FFAB91'}:{f:'#FBE9E7',s:'#D84315',t:'#3E0A00'},
  edge: isDark?'rgba(180,178,169,0.18)':'rgba(100,100,100,0.18)',
  edgeHi:'#378ADD',edgeUp:'#1B8A5A',edgeDn:'#E65100',
  layerBg:isDark?'rgba(180,178,169,0.05)':'rgba(100,100,100,0.04)',
  layerTx:isDark?'rgba(180,178,169,0.4)':'rgba(60,60,60,0.5)',
};

function _buildGraph(){
  const _hm={},_srcFlg={},_tgtFlg={};
  for(const r of RECS){
    _hm[r.src]=r.src_hop;_hm[r.tgt]=r.tgt_hop;
    if(r.is_src)_srcFlg[r.src]=1;
    if(r.tgt_hop===0)_tgtFlg[r.tgt]=1;
  }
  MAX_HOP=Math.max(...Object.values(_hm),0);
  // Deduplicate: same (src,tgt) via different jobs → one edge, merged labels
  const _edgeMap=new Map();
  for(const r of RECS){
    const k=r.src+'→'+r.tgt;
    if(!_edgeMap.has(k))_edgeMap.set(k,{s:r.src,t:r.tgt,jobs:new Set(),dfs:new Set(),wfs:new Set()});
    const e=_edgeMap.get(k);
    r.job.split(';').map(x=>x.trim()).filter(Boolean).forEach(j=>e.jobs.add(j));
    r.df.split(';').map(x=>x.trim()).filter(Boolean).forEach(d=>e.dfs.add(d));
    r.wf.split(';').map(x=>x.trim()).filter(Boolean).forEach(w=>e.wfs.add(w));
  }
  NODE_TYPE={};
  for(const nm of Object.keys(_hm)){
    if(_tgtFlg[nm])NODE_TYPE[nm]='tgt';
    else if(nm.startsWith('SAP.'))NODE_TYPE[nm]='sap';
    else if(nm.startsWith('FILE:'))NODE_TYPE[nm]='file';
    else if(nm.startsWith('EXCEL:'))NODE_TYPE[nm]='excel';
    else if(nm.startsWith('SQL:'))NODE_TYPE[nm]='sql';
    else if(_srcFlg[nm])NODE_TYPE[nm]='src';
    else NODE_TYPE[nm]='xfm';
  }
  // SQL cluster: members come from ds.sql_members (built by Python Phase E1).
  // Each SQL:xxx node's physical DWH tables are pre-computed and passed in JSON.
  // Double-click the SQL node to expand/collapse the cluster.
  SQL_MEMBERS={};
  const _curDs=DATASETS[currentTab]||{};
  for(const [sn,mems] of Object.entries(_curDs.sql_members||{})){
    // Skip SQL nodes not present in this graph's hop-map (filtered by routing)
    if(_hm[sn]==null)continue;
    // Only include members that actually appear in the graph (present in _hm)
    const valid=mems.filter(m=>{const mu=m.toUpperCase();return _hm[mu]!=null||_hm[m]!=null;});
    if(valid.length>0){SQL_MEMBERS[sn]=valid;if(!expandedSqlNodes.has(sn))collapsedSqlNodes.add(sn);}
  }
  const _hiddenNodes=new Set();
  for(const[sn,mems]of Object.entries(SQL_MEMBERS)){
    if(collapsedSqlNodes.has(sn))mems.forEach(m=>_hiddenNodes.add(m));
  }
  // Build member→SQL-cluster map for collapsed clusters so broken edges can be
  // redirected: upstream→hidden-member becomes upstream→SQL-cluster-node.
  const _mbrToSql={};
  for(const[sn,mems]of Object.entries(SQL_MEMBERS)){
    if(collapsedSqlNodes.has(sn))mems.forEach(m=>{_mbrToSql[m]=sn;});
  }
  // Build EDGES: skip edges whose source is hidden; redirect edges whose target
  // is a hidden member to the SQL cluster node instead of dropping them.
  // This fixes the visual gap where upstream nodes (Hop 2+) lose their connections
  // when a SQL cluster is collapsed.
  const _edgeDedup=new Map();
  for(const e of _edgeMap.values()){
    if(_hiddenNodes.has(e.s))continue;                        // source hidden → drop
    const t=_hiddenNodes.has(e.t)?(_mbrToSql[e.t]||null):e.t; // redirect or null
    if(!t)continue;                                            // no redirect → drop
    const k=e.s+'→'+t;
    if(!_edgeDedup.has(k))_edgeDedup.set(k,{s:e.s,t,jobs:new Set(),dfs:new Set(),wfs:new Set()});
    const m=_edgeDedup.get(k);
    e.jobs.forEach(j=>m.jobs.add(j));e.dfs.forEach(d=>m.dfs.add(d));e.wfs.forEach(w=>m.wfs.add(w));
  }
  EDGES=Array.from(_edgeDedup.values()).map(e=>[e.s,e.t,[...e.jobs].join(';'),[...e.dfs].join(';'),[...e.wfs].join(';')]);
  LAYERS=Array.from({length:MAX_HOP+1},()=>[]);
  for(const[nm,h]of Object.entries(_hm)){
    if(_hiddenNodes.has(nm))continue;
    const li=MAX_HOP-h;
    if(!LAYERS[li].includes(nm))LAYERS[li].push(nm);
  }
  nodeJobs={};nodeDFs={};nodeWFs={};
  for(const[s,t,job,df,wf]of EDGES){
    for(const nd of[s,t]){
      if(!nodeJobs[nd])nodeJobs[nd]=new Set();
      if(!nodeDFs[nd])nodeDFs[nd]=new Set();
      if(!nodeWFs[nd])nodeWFs[nd]=new Set();
      job.split(';').map(x=>x.trim()).filter(Boolean).forEach(j=>nodeJobs[nd].add(j));
      df.split(';').map(x=>x.trim()).filter(Boolean).forEach(d=>nodeDFs[nd].add(d));
      wf.split(';').map(x=>x.trim()).filter(Boolean).forEach(w=>nodeWFs[nd].add(w));
    }
  }
  multiJob=new Set();
  for(const[nd,js]of Object.entries(nodeJobs)){if(js.size>1)multiJob.add(nd);}
  positions={};layerXs=[];
  let lx=PAD;
  for(let li=0;li<LAYERS.length;li++){layerXs.push(lx);lx+=NW+HGAP;}
  totalW=lx-HGAP+PAD;
  for(let li=0;li<LAYERS.length;li++){
    LAYERS[li].forEach((n,ni)=>{positions[n]={x:layerXs[li],y:PAD+LHDR+ni*(NH+VGAP)};});
  }
  const maxNodes=Math.max(...LAYERS.map(l=>l.length),0);
  totalH=PAD+LHDR+maxNodes*(NH+VGAP)+PAD;
  canvas.width=totalW;canvas.height=totalH;
  // ── Barycenter layer ordering (3 forward + backward passes) ─────────────
  // Sorts nodes within each layer by the average y of their connected neighbors,
  // minimising the number of edge crossings (Sugiyama step 3).
  function _repos(){
    for(let li=0;li<LAYERS.length;li++)
      LAYERS[li].forEach((n,ni)=>{positions[n]={x:layerXs[li],y:PAD+LHDR+ni*(NH+VGAP)};});
  }
  for(let pass=0;pass<4;pass++){
    // Forward sweep: sort by avg y of right-neighbours
    for(let li=0;li<LAYERS.length-1;li++){
      const bary={};
      for(const n of LAYERS[li]){
        const ys=EDGES.filter(([s])=>s===n).map(([,t])=>positions[t]?positions[t].y+NH/2:null).filter(v=>v!==null);
        bary[n]=ys.length?ys.reduce((a,b)=>a+b,0)/ys.length:positions[n].y;
      }
      LAYERS[li].sort((a,b)=>bary[a]-bary[b]);
      _repos();
    }
    // Backward sweep: sort by avg y of left-neighbours
    for(let li=LAYERS.length-1;li>0;li--){
      const bary={};
      for(const n of LAYERS[li]){
        const ys=EDGES.filter(([s,t])=>t===n).map(([s])=>positions[s]?positions[s].y+NH/2:null).filter(v=>v!==null);
        bary[n]=ys.length?ys.reduce((a,b)=>a+b,0)/ys.length:positions[n].y;
      }
      LAYERS[li].sort((a,b)=>bary[a]-bary[b]);
      _repos();
    }
  }
  // ── Port-slot assignment: sort by neighbour y so departure order matches arrival order ──
  _outEdges={};_inEdges={};
  for(let ei=0;ei<EDGES.length;ei++){
    const[s,t]=EDGES[ei];
    (_outEdges[s]||(_outEdges[s]=[])).push(ei);
    (_inEdges[t] ||(_inEdges[t] =[])).push(ei);
  }
  for(const eis of Object.values(_outEdges)){
    eis.sort((a,b)=>{
      const ya=positions[EDGES[a][1]]?positions[EDGES[a][1]].y:0;
      const yb=positions[EDGES[b][1]]?positions[EDGES[b][1]].y:0;
      return ya-yb;
    });
  }
  for(const eis of Object.values(_inEdges)){
    eis.sort((a,b)=>{
      const ya=positions[EDGES[a][0]]?positions[EDGES[a][0]].y:0;
      const yb=positions[EDGES[b][0]]?positions[EDGES[b][0]].y:0;
      return ya-yb;
    });
  }
  edgeSlots=EDGES.map((_,ei)=>{
    const[s,t]=EDGES[ei];
    return[_outEdges[s].indexOf(ei),_inEdges[t].indexOf(ei)];
  });
}

function initTab(i){
  currentTab=i;
  hov=null;sel=null;spotDir=null;searchStr='';
  expandedSqlNodes=new Set();collapsedSqlNodes=new Set();
  fieldSearchStr='';fieldSearchNodes=new Set();fieldSearchEdges=new Map();
  _clearPath();
  _traceMode=false;_traceNodes=new Set();_traceEdgeKeys=new Set();
  _cpAllDfs=[];_cpData=null;
  document.getElementById('colPanel').classList.remove('open');
  document.getElementById('cpDfPicker').style.display='none';
  document.getElementById('cpNormalTabs').style.display='';
  document.getElementById('cpTraceHdr').classList.remove('open');
  RECS=DATASETS[i].recs;TGT=DATASETS[i].tgt;
  _buildGraph();
  document.getElementById('htbl').textContent=DATASETS[i].label||TGT;
  document.getElementById('search').value='';
  document.getElementById('fieldSearch').value='';
  document.getElementById('spotBadge').style.display='none';
  document.getElementById('detail').innerHTML='Click any table node to see job names, dataflows, and connections.';
  document.querySelectorAll('.tab-btn').forEach((b,j)=>b.classList.toggle('active',j===i));
  draw();
}
(function(){
  const strip=document.getElementById('tabStrip');
  if(DATASETS.length<=1){strip.classList.add('hidden');document.getElementById('btnExpAll').style.display='none';return;}
  DATASETS.forEach((ds,i)=>{
    const b=document.createElement('button');
    b.className='tab-btn'+(i===0?' active':'');
    b.textContent=ds.label?ds.label.slice(0,30):ds.tgt.split('.').pop().slice(0,30);
    b.title=ds.targets?ds.targets.join('\n'):ds.tgt;
    b.onclick=()=>initTab(i);
    strip.appendChild(b);
  });
})();

function nodeColor(id){if(multiJob.has(id))return C.multi;return C[NODE_TYPE[id]||'xfm']||C.xfm;}
function shortLabel(n){
  if(n.startsWith('FILE:')){const f=n.split('/').pop();return f.length>20?f.slice(0,18)+'…':f;}
  if(n.startsWith('EXCEL:')){const parts=n.slice(6).split('/');const sheet=parts.pop();return sheet.length>20?sheet.slice(0,18)+'…':sheet;}
  if(n.startsWith('SAP.')){const t=n.slice(4);return t.length>20?t.slice(0,18)+'…':t;}
  if(n.startsWith('SQL:')){
    // Node key is compound "SQL:df_name|tf_name" — strip df_name| for display.
    const raw=n.slice(4);
    const base=raw.includes('|')?raw.split('|').pop():raw;
    if(SQL_MEMBERS[n]&&collapsedSqlNodes.has(n))return(base.length>12?base.slice(0,10)+'…':base)+' [▶'+SQL_MEMBERS[n].length+']';
    return base.length>18?'SQL:'+base.slice(0,15)+'…':'SQL:'+base;
  }
  const p=n.split('.');const nm=p[p.length-1];return nm.length>20?nm.slice(0,18)+'…':nm;
}
function rr(x,y,w,h,r){
  ctx.beginPath();
  ctx.moveTo(x+r,y);ctx.lineTo(x+w-r,y);ctx.arcTo(x+w,y,x+w,y+r,r);
  ctx.lineTo(x+w,y+h-r);ctx.arcTo(x+w,y+h,x+w-r,y+h,r);
  ctx.lineTo(x+r,y+h);ctx.arcTo(x,y+h,x,y+h-r,r);
  ctx.lineTo(x,y+r);ctx.arcTo(x,y,x+r,y,r);
  ctx.closePath();
}
function computePath(id){
  upNodes=new Set([id]);upEdgeKeys=new Set();
  dnNodes=new Set([id]);dnEdgeKeys=new Set();
  const upQ=[id];
  while(upQ.length){const n=upQ.pop();for(const[s,t]of EDGES){if(t===n){upEdgeKeys.add(s+'→'+t);if(!upNodes.has(s)){upNodes.add(s);upQ.push(s);}}}}
  const dnQ=[id];
  while(dnQ.length){const n=dnQ.pop();for(const[s,t]of EDGES){if(s===n){dnEdgeKeys.add(s+'→'+t);if(!dnNodes.has(t)){dnNodes.add(t);dnQ.push(t);}}}}
  _applyDir();
}
function _applyDir(){
  if(spotDir==='up'){pathNodes=new Set(upNodes);pathEdgeKeys=new Set(upEdgeKeys);}
  else if(spotDir==='dn'){pathNodes=new Set(dnNodes);pathEdgeKeys=new Set(dnEdgeKeys);}
  else{pathNodes=new Set([...upNodes,...dnNodes]);pathEdgeKeys=new Set([...upEdgeKeys,...dnEdgeKeys]);}
}
function _clearPath(){
  pathNodes=new Set();pathEdgeKeys=new Set();
  upNodes=new Set();upEdgeKeys=new Set();
  dnNodes=new Set();dnEdgeKeys=new Set();
}
function draw(){
  ctx.clearRect(0,0,canvas.width,canvas.height);
  for(let li=0;li<LAYERS.length;li++){
    const lxb=layerXs[li];
    ctx.fillStyle=C.layerBg;ctx.fillRect(lxb-HGAP/2,0,NW+HGAP,totalH);
    const hop=MAX_HOP-li;
    const lbl=li===LAYERS.length-1?'🎯 Target':(li===0?'🌿 Sources':'Hop '+hop);
    ctx.fillStyle=C.layerTx;ctx.font='10px system-ui,sans-serif';ctx.textAlign='center';
    ctx.fillText(lbl,lxb+NW/2,PAD+10);
  }
  const hasSel=sel!==null,hasSearch=searchStr.length>0,isSpot=spotDir!==null,isTrace=_traceMode;
  // Helper: compute geometry + style for one edge index
  function _edgeGeom(ei){
    const[s,t]=EDGES[ei];
    const sp=positions[s],tp=positions[t];if(!sp||!tp)return null;
    const edgeKey=s+'→'+t;
    const onPath=hasSel&&pathEdgeKeys.has(edgeKey);
    const hovHi=!hasSel&&!hasSearch&&(hov===s||hov===t);
    const searchHi=hasSearch&&(s.includes(searchStr)||t.includes(searchStr));
    const _fek=(s+'→'+t).toUpperCase();
    const fieldHi=fieldSearchStr.length>0&&fieldSearchEdges.has(_fek);
    const srcCol=(C[NODE_TYPE[s]||'xfm']||C.xfm).s;
    let ecol,elw,ealpha,isHi=false,tracedEdge=false;
    if(isTrace){
      if(_traceEdgeKeys.has(edgeKey)){ecol='#7C3AED';elw=2.5;ealpha=1;isHi=true;tracedEdge=true;}
      else{ecol=C.edge;elw=0.7;ealpha=0.04;}
    }
    else if(onPath){
      if(isSpot)ecol=upEdgeKeys.has(edgeKey)?C.edgeUp:C.edgeDn;
      else ecol=C.edgeHi;
      elw=2.5;ealpha=1;isHi=true;
    }
    else if(hovHi){ecol=C.edgeHi;elw=2;ealpha=0.9;isHi=true;}
    else if(searchHi){ecol='#F9A825';elw=1.5;ealpha=0.9;isHi=true;}
    else if(fieldHi){ecol='#7C3AED';elw=2;ealpha=0.9;isHi=true;}
    else if(hasSel){ecol=C.edge;elw=0.7;ealpha=isSpot?0.03:0.07;}
    else if(hasSearch){ecol=C.edge;elw=0.7;ealpha=0.06;}
    else{ecol=srcCol;elw=0.7;ealpha=0.28;}
    const[oSlot,iSlot]=edgeSlots[ei]||[0,0];
    const oC=(_outEdges[s]||[]).length,iC=(_inEdges[t]||[]).length;
    const x1=sp.x+NW,y1=sp.y+NH*(oSlot+1)/(oC+1);
    const x2=tp.x,   y2=tp.y+NH*(iSlot+1)/(iC+1);
    return{s,t,edgeKey,x1,y1,x2,y2,ecol,elw,ealpha,isHi,onPath,isSpot,upEdgeKeys,tracedEdge,fieldHi,_fek};
  }
  function _drawEdge(g){
    const{x1,y1,x2,y2,ecol,elw,ealpha,isHi,onPath,isSpot}=g;
    const mx=(x1+x2)/2;
    // Gradient on highlighted edges: light-at-source → dark-at-target shows direction at crossings
    let strokeCol=ecol,fillCol=ecol;
    if(isHi){
      const grad=ctx.createLinearGradient(x1,y1,x2,y2);
      let c0,c1;
      if(g.tracedEdge){c0='#D1C4E9';c1='#4527A0';}                   // field trace: pale lavender→deep violet
      else if(onPath&&isSpot){
        if(g.upEdgeKeys.has(g.edgeKey)){c0='#B7E4CB';c1='#0A5C3A';}  // upstream: pale→deep green
        else                            {c0='#FDDBA8';c1='#BF360C';}  // downstream: pale amber→deep red-orange
      } else {c0='#A9D4F5';c1='#1A5276';}                             // hover/both: pale→deep blue
      grad.addColorStop(0,c0);grad.addColorStop(1,c1);
      strokeCol=grad;fillCol=c1;
    }
    ctx.strokeStyle=strokeCol;ctx.lineWidth=elw;ctx.globalAlpha=ealpha;
    ctx.beginPath();ctx.moveTo(x1,y1);ctx.bezierCurveTo(mx,y1,mx,y2,x2,y2);ctx.stroke();
    const aLen=elw>=2?13:9,aW=elw>=2?5.5:4;
    ctx.fillStyle=fillCol;
    ctx.beginPath();ctx.moveTo(x2,y2);ctx.lineTo(x2-aLen,y2-aW);ctx.lineTo(x2-aLen,y2+aW);ctx.closePath();ctx.fill();
    if(Math.abs(y2-y1)>NH){
      const mAng=Math.atan2(2*(y2-y1),x2-x1);
      const maL=aLen*0.75,maW=aW*0.7,bx=mx,by=(y1+y2)/2;
      const mbx=bx-maL*Math.cos(mAng),mby=by-maL*Math.sin(mAng);
      ctx.beginPath();ctx.moveTo(bx,by);
      ctx.lineTo(mbx-maW*Math.sin(mAng),mby+maW*Math.cos(mAng));
      ctx.lineTo(mbx+maW*Math.sin(mAng),mby-maW*Math.cos(mAng));
      ctx.closePath();ctx.fill();
    }
    ctx.globalAlpha=1;
  }
  // Two-pass draw: dimmed/background edges first, highlighted on top so they're never occluded
  const hiEdges=[];
  for(let ei=0;ei<EDGES.length;ei++){
    const g=_edgeGeom(ei);if(!g)continue;
    if(g.isHi)hiEdges.push(g);else _drawEdge(g);
  }
  for(const g of hiEdges)_drawEdge(g);
  // Number badges on field-search highlighted edges (hop-ordered sequence)
  if(fieldSearchStr){
    for(const g of hiEdges){
      if(!g.fieldHi)continue;
      const seq=fieldSearchEdges.get(g._fek);if(!seq)continue;
      const bx=g.x1*0.45+g.x2*0.55;const by=g.y1*0.45+g.y2*0.55;
      ctx.globalAlpha=1;
      ctx.fillStyle='#7C3AED';ctx.beginPath();ctx.arc(bx,by,9,0,Math.PI*2);ctx.fill();
      ctx.fillStyle='#fff';ctx.font='bold 8px Consolas,monospace';
      ctx.textAlign='center';ctx.textBaseline='middle';
      ctx.fillText(String(seq),bx,by);
    }
  }
  for(const[id,pos]of Object.entries(positions)){
    const c=nodeColor(id),active=(hov===id||sel===id);
    const onPath=hasSel&&pathNodes.has(id);
    const searchMatch=hasSearch&&id.includes(searchStr);
    const nodeDimmed=isTrace?!_traceNodes.has(id):((hasSel&&!onPath)||(hasSearch&&!searchMatch));
    const nodeAlpha=nodeDimmed?(isSpot?0.04:0.15):1;
    ctx.globalAlpha=nodeAlpha;
    ctx.fillStyle=c.f;ctx.strokeStyle=c.s;ctx.lineWidth=active?2:0.8;
    ctx.shadowColor=active?c.s+'66':'transparent';ctx.shadowBlur=active?7:0;
    rr(pos.x,pos.y,NW,NH,4);ctx.fill();ctx.stroke();
    ctx.shadowBlur=0;ctx.globalAlpha=1;
    if(isSpot&&hasSel&&onPath){
      const inUp=upNodes.has(id),inDn=dnNodes.has(id),isPivot=(id===sel);
      const ringCol=isPivot?'#FFD600':inUp&&inDn?'#FFD600':inUp?C.edgeUp:C.edgeDn;
      ctx.strokeStyle=ringCol;ctx.lineWidth=isPivot?2.5:1.5;ctx.globalAlpha=0.9;
      rr(pos.x-1,pos.y-1,NW+2,NH+2,5);ctx.stroke();ctx.globalAlpha=1;
    }
    if(isTrace&&_traceNodes.has(id)){
      ctx.strokeStyle='#7C3AED';ctx.lineWidth=1.5;ctx.globalAlpha=0.8;
      rr(pos.x-1,pos.y-1,NW+2,NH+2,5);ctx.stroke();ctx.globalAlpha=1;
    }
    if(searchMatch&&!active){
      ctx.strokeStyle='#F9A825';ctx.lineWidth=2;ctx.globalAlpha=0.9;
      rr(pos.x-1,pos.y-1,NW+2,NH+2,5);ctx.stroke();ctx.globalAlpha=1;
    }
    if(fieldSearchNodes.size&&fieldSearchNodes.has(id.toUpperCase())){
      ctx.strokeStyle='#7C3AED';ctx.lineWidth=2;ctx.globalAlpha=0.85;
      rr(pos.x-2,pos.y-2,NW+4,NH+4,6);ctx.stroke();ctx.globalAlpha=1;
    }
    ctx.globalAlpha=nodeAlpha;
    ctx.fillStyle=nodeDimmed?'rgba(150,148,140,0.35)':c.t;
    ctx.font=(active?'500 ':'')+'11px Consolas,monospace';
    ctx.textAlign='center';ctx.textBaseline='middle';
    ctx.fillText(shortLabel(id),pos.x+NW/2,pos.y+NH/2);
    ctx.globalAlpha=1;
  }
}
function hit(mx,my){for(const[id,p]of Object.entries(positions))if(mx>=p.x&&mx<=p.x+NW&&my>=p.y&&my<=p.y+NH)return id;return null;}
canvas.addEventListener('mousemove',function(e){
  const r=canvas.getBoundingClientRect(),sc=canvas.width/r.width;
  const h=hit((e.clientX-r.left)*sc,(e.clientY-r.top)*sc);
  if(h!==hov){hov=h;canvas.style.cursor=h?'pointer':'default';draw();}
});
canvas.addEventListener('click',function(e){
  dismissCtxMenu();
  const r=canvas.getBoundingClientRect(),sc=canvas.width/r.width;
  const cx=(e.clientX-r.left)*sc,cy=(e.clientY-r.top)*sc;
  const h=hit(cx,cy);
  if(!h){
    const eg=hitEdge(cx,cy);
    if(eg){closeSqlPanel();openColPanel(eg.s,eg.t);return;}
  }
  closeColPanel();
  sel=(h===sel)?null:h;spotDir=null;
  if(sel){computePath(sel);}else{_clearPath();}
  document.getElementById('spotBadge').style.display='none';
  draw();
  if(sel){
    if(NODE_TYPE[sel]==='sql'){openSqlPanel(sel);}
    else{closeSqlPanel();}
    showDetail(sel);
  }else{
    closeSqlPanel();
    document.getElementById('detail').innerHTML='Click a node for path &amp; details &middot; Click an edge for column-level lineage &middot; Right-click for spotlight.';
  }
});
canvas.addEventListener('contextmenu',function(e){
  e.preventDefault();
  const r=canvas.getBoundingClientRect(),sc=canvas.width/r.width;
  const h=hit((e.clientX-r.left)*sc,(e.clientY-r.top)*sc);
  if(h)showCtxMenu(e.clientX,e.clientY,h);else dismissCtxMenu();
});
canvas.addEventListener('dblclick',function(e){
  const r=canvas.getBoundingClientRect(),sc=canvas.width/r.width;
  const h=hit((e.clientX-r.left)*sc,(e.clientY-r.top)*sc);
  if(h&&SQL_MEMBERS[h]){
    if(collapsedSqlNodes.has(h)){collapsedSqlNodes.delete(h);expandedSqlNodes.add(h);}
    else{collapsedSqlNodes.add(h);expandedSqlNodes.delete(h);}
    _buildGraph();draw();
    const isCollapsed=collapsedSqlNodes.has(h);
    const _qt=((DATASETS[currentTab]||{}).sql_queries||{})[h]||'';
    const _hRaw=h.slice(4);const _hDisp=_hRaw.includes('|')?_hRaw.split('|').pop():_hRaw;
    document.getElementById('detail').innerHTML='SQL cluster <b>'+_esc(_hDisp)+'</b>: '
      +(isCollapsed?'collapsed &mdash; '+SQL_MEMBERS[h].length+' source table(s) hidden. Double-click again to expand.'
                  :'expanded &mdash; '+SQL_MEMBERS[h].length+' source table(s) shown. Double-click again to collapse.')
      +(_qt?'<br><span style="color:#888780;font-size:10px">SQL query:</span><br><code style="font-size:10px;display:block;white-space:pre-wrap;background:#FBE9E7;padding:6px 8px;border-radius:4px;margin-top:2px;color:#3E0A00;max-height:380px;overflow-y:auto;line-height:1.5;font-family:Consolas,monospace">'+_esc(_qt)+'</code>':'');
  }
});
function badge(txt,col){return '<span class="tag" style="background:'+col+'22;color:'+col+';border:1px solid '+col+'55">'+txt+'</span>';}
function showDetail(id){
  const c=nodeColor(id);
  const jobs=[...(nodeJobs[id]||[])],dfs=[...(nodeDFs[id]||[])],wfs=[...(nodeWFs[id]||[])];
  const ins=EDGES.filter(e=>e[0]===id),outs=EDGES.filter(e=>e[1]===id);
  const tp=NODE_TYPE[id]||'xfm';
  const typeLabel=tp==='tgt'?'Target table':tp==='src'?'DB Source':tp==='sap'?'SAP Source':tp==='file'?'Flat File':tp==='excel'?'Excel Workbook':tp==='sql'?'SQL Transform':multiJob.has(id)?'Multi-job intermediate':'Intermediate';
  let html='<b style="color:'+c.s+'">'+id+'</b>&nbsp;<span style="font-size:10px;color:#888780">'+typeLabel+'</span>';
  if(multiJob.has(id))html+=' <span style="font-size:10px;color:#D4537E;font-weight:500">&#9888; shared by '+jobs.length+' jobs</span>';
  html+='<br>';
  if(jobs.length)html+='<span style="color:#888780">Jobs:</span> '+jobs.map(j=>badge(j,'#378ADD')).join(' ')+'<br>';
  if(wfs.length)html+='<span style="color:#888780">Workflows:</span> '+wfs.map(w=>badge(w,'#8E24AA')).join(' ')+'<br>';
  if(dfs.length){const show=dfs.slice(0,6);html+='<span style="color:#888780">Dataflows:</span> '+show.map(d=>badge(d,'#1D9E75')).join(' ')+(dfs.length>6?' <span style="color:#888780">+'+(dfs.length-6)+' more</span>':'')+'<br>';}
  // For SQL transform nodes: extract source tables + SQL text from cols data
  if(tp==='sql'){
    // id is compound "SQL:df_name|tf_name"; extract just tf_name for df lookup + display.
    const raw=id.slice(4);
    const dfName=raw.includes('|')?raw.split('|').pop():raw;
    const tgtNodes=[...new Set(outs.map(e=>e[1]))];
    const ds=DATASETS[currentTab];
    let sqlSrcs=[],sqlText='';
    for(const tgtId of tgtNodes){
      const colEntry=ds&&ds.cols&&(ds.cols[tgtId]||ds.cols[tgtId.split('.').pop()]||ds.cols[tgtId.toUpperCase()])||null;
      if(!colEntry)continue;
      for(const dfEntry of(colEntry.dataflows||[])){
        if(dfEntry.df!==dfName)continue;
        for(const m of(dfEntry.mappings||[])){
          if((m.ft||'').toUpperCase()==='SQL_CUSTOM'){
            if(m.src_obj)sqlSrcs.push(...m.src_obj.split(',').map(s=>s.trim()).filter(Boolean));
            if(!sqlText&&m.f)sqlText=m.f;
          }
        }
      }
    }
    const uniqSrcs=[...new Set(sqlSrcs)];
    // Phase E1: fall back to sql_queries dict for transforms captured at BFS level
    if(!sqlText&&ds&&ds.sql_queries&&ds.sql_queries[id])sqlText=ds.sql_queries[id];
    // Show SQL cluster members as source table badges if not already populated
    if(!uniqSrcs.length&&SQL_MEMBERS[id])SQL_MEMBERS[id].forEach(m=>{if(!uniqSrcs.includes(m))uniqSrcs.push(m);});
    if(uniqSrcs.length)html+='<span style="color:#888780">Source tables:</span> '+uniqSrcs.map(s=>badge(s.split('.').pop(),'#D84315')).join(' ')+'<br>';
    if(sqlText){html+='<span style="color:#888780">SQL query:</span><br><code style="font-size:10px;display:block;white-space:pre-wrap;background:#FBE9E7;padding:6px 8px;border-radius:4px;margin-top:2px;color:#3E0A00;max-height:380px;overflow-y:auto;line-height:1.5;font-family:Consolas,monospace">'+_esc(sqlText)+'</code>';}
  }
  const upNds=[...new Set(ins.map(e=>e[0]))],dnNds=[...new Set(outs.map(e=>e[1]))];
  if(upNds.length)html+='<span style="color:#888780">Receives from ('+upNds.length+'):</span> '+upNds.map(n=>'<b>'+shortLabel(n)+'</b>').join(', ')+'<br>';
  if(dnNds.length)html+='<span style="color:#888780">Feeds into ('+dnNds.length+'):</span> '+dnNds.map(n=>'<b>'+shortLabel(n)+'</b>').join(', ');
  document.getElementById('detail').innerHTML=html;
}
function showCtxMenu(x,y,id){
  dismissCtxMenu();_ctxTarget=id;
  const c=nodeColor(id),tp=NODE_TYPE[id]||'xfm';
  const typeLabel=tp==='tgt'?'Target':tp==='src'?'DB Source':tp==='sap'?'SAP Source':tp==='file'?'Flat File':tp==='excel'?'Excel':tp==='sql'?'SQL Transform':multiJob.has(id)?'Multi-job':'Intermediate';
  const menu=document.createElement('div');menu.className='ctx-menu';menu.id='ctxMenu';
  menu.innerHTML=
    '<div class="ctx-hdr"><span style="color:'+c.s+';font-weight:500">'+shortLabel(id)+'</span>&nbsp;&middot;&nbsp;'+typeLabel+'</div>'+
    '<div class="ctx-item" id="cSpotUp"><span style="color:'+C.edgeUp+'">&#11014;</span> Upstream spotlight</div>'+
    '<div class="ctx-item" id="cSpotDn"><span style="color:'+C.edgeDn+'">&#11015;</span> Downstream spotlight</div>'+
    '<div class="ctx-item" id="cSpotBoth">&#8597; Full spotlight (both)</div>'+
    '<div class="ctx-sep"></div>'+
    '<div class="ctx-item" id="cPath">&#128204; Highlight path (dim)</div>'+
    '<div class="ctx-item" id="cCopy">&#128203; Copy full name</div>'+
    '<div class="ctx-item" id="cInfo">&#8505;&#65039; Details</div>';
  document.body.appendChild(menu);
  menu.querySelector('#cSpotUp').onclick=()=>doSpotlight(_ctxTarget,'up');
  menu.querySelector('#cSpotDn').onclick=()=>doSpotlight(_ctxTarget,'dn');
  menu.querySelector('#cSpotBoth').onclick=()=>doSpotlight(_ctxTarget,'both');
  menu.querySelector('#cPath').onclick=()=>{dismissCtxMenu();doSelect(_ctxTarget);};
  menu.querySelector('#cCopy').onclick=()=>{navigator.clipboard&&navigator.clipboard.writeText(_ctxTarget);dismissCtxMenu();};
  menu.querySelector('#cInfo').onclick=()=>{dismissCtxMenu();doSelect(_ctxTarget);};
  menu.style.left=x+'px';menu.style.top=y+'px';
  const rect=menu.getBoundingClientRect();
  if(rect.right>window.innerWidth-8)menu.style.left=(x-rect.width)+'px';
  if(rect.bottom>window.innerHeight-8)menu.style.top=(y-rect.height)+'px';
}
function dismissCtxMenu(){const m=document.getElementById('ctxMenu');if(m)m.remove();}
function doSpotlight(id,dir){
  dismissCtxMenu();sel=id;spotDir=dir;computePath(id);
  const bdg=document.getElementById('spotBadge'),lbl=document.getElementById('spotLabel');
  const col=dir==='up'?C.edgeUp:dir==='dn'?C.edgeDn:'#378ADD';
  const icon=dir==='up'?'&#11014; Upstream':dir==='dn'?'&#11015; Downstream':'&#8597; Full path';
  lbl.innerHTML=icon+' spotlight';
  bdg.style.background=col+'18';bdg.style.borderColor=col+'55';
  bdg.style.color=col;bdg.querySelector('button').style.color=col;
  bdg.style.display='inline-flex';draw();showDetail(id);
}
function clearSpotlight(){
  sel=null;spotDir=null;_clearPath();
  document.getElementById('spotBadge').style.display='none';
  draw();document.getElementById('detail').innerHTML='Click a node for path &amp; details &middot; Right-click for spotlight.';
}
function doSelect(id){
  sel=(sel===id)?null:id;spotDir=null;
  if(sel){computePath(sel);}else{_clearPath();}
  document.getElementById('spotBadge').style.display='none';
  draw();
  if(sel)showDetail(sel);
  else document.getElementById('detail').innerHTML='Click a node for path &amp; details &middot; Right-click for spotlight.';
}
document.getElementById('search').addEventListener('input',function(e){searchStr=e.target.value.trim().toUpperCase();draw();});
document.getElementById('fieldSearch').addEventListener('input',function(e){
  fieldSearchStr=(e.target.value||'').trim();
  fieldSearchNodes=new Set();fieldSearchEdges=new Map();
  if(!fieldSearchStr){draw();return;}
  const q=fieldSearchStr.toUpperCase();
  const ds=DATASETS[currentTab];
  let seq=1;
  for(const[tbl,tblData]of Object.entries(ds.cols||{})){
    for(const dfEntry of(tblData.dataflows||[])){
      for(const m of(dfEntry.mappings||[])){
        if((m.sc||'').toUpperCase().includes(q)||(m.tc||'').toUpperCase().includes(q)){
          const srcObj=m.src_obj||tbl;
          fieldSearchNodes.add(tbl.toUpperCase());
          fieldSearchNodes.add(srcObj.toUpperCase());
          const ek=(srcObj+'→'+tbl).toUpperCase();
          if(!fieldSearchEdges.has(ek))fieldSearchEdges.set(ek,seq++);
        }
      }
    }
  }
  draw();
});
document.addEventListener('click',function(e){if(!e.target.closest('#ctxMenu'))dismissCtxMenu();});
document.addEventListener('keydown',function(e){
  if(e.key==='Escape'){
    dismissCtxMenu();
    const inp=document.getElementById('search');
    if(searchStr){inp.value='';searchStr='';draw();}
    else if(spotDir!==null){clearSpotlight();}
    else if(sel){doSelect(null);}
  }
  if((e.key==='f'||e.key==='F')&&(e.ctrlKey||e.metaKey)){e.preventDefault();document.getElementById('search').focus();}
  if((e.key==='g'||e.key==='G')&&(e.ctrlKey||e.metaKey)){e.preventDefault();document.getElementById('fieldSearch').focus();}
  if(e.key==='Escape'&&fieldSearchStr){fieldSearchStr='';fieldSearchNodes=new Set();fieldSearchEdges=new Map();document.getElementById('fieldSearch').value='';draw();}
});
function _rows(recs){
  return recs.map(r=>({Target:r.tgt,Target_Hop:r.tgt_hop,Source:r.src,Source_Hop:r.src_hop,
    Is_Terminal_Source:r.is_src?'Yes':'No',Job:r.job,Workflow:r.wf,Dataflow:r.df}));
}
function _summary(ds){
  const nodes={};for(const r of ds.recs){nodes[r.src]=r.src_hop;nodes[r.tgt]=r.tgt_hop;}
  const terms=[...new Set(ds.recs.filter(r=>r.is_src).map(r=>r.src))];
  return{Target:ds.tgt,Records:ds.recs.length,Max_Hops:Math.max(...Object.values(nodes),0),
    Terminal_Sources:terms.length,SAP_Sources:terms.filter(n=>n.startsWith('SAP.')).length,
    File_Sources:terms.filter(n=>n.startsWith('FILE:')).length,
    Excel_Sources:terms.filter(n=>n.startsWith('EXCEL:')).length};
}
function _csvDl(rows,fname){
  if(!rows.length)return;
  const ks=Object.keys(rows[0]);
  const body=[ks,...rows.map(r=>ks.map(k=>'"'+String(r[k]!=null?r[k]:'').replace(/"/g,'""')+'"'))].map(r=>r.join(',')).join('\n');
  const a=document.createElement('a');a.href='data:text/csv;charset=utf-8,﻿'+encodeURIComponent(body);a.download=fname;a.click();
}
function _xlsxDl(datasets,fname){
  const wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb,XLSX.utils.json_to_sheet(datasets.map(_summary)),'Summary');
  datasets.forEach(ds=>{
    const safe=ds.tgt.replace(/[^A-Za-z0-9_]/g,'_').slice(0,31);
    XLSX.utils.book_append_sheet(wb,XLSX.utils.json_to_sheet(_rows(ds.recs)),safe);
  });
  XLSX.writeFile(wb,fname);
}
function exportCurrentTab(){
  const ds=DATASETS[currentTab],safe=ds.tgt.replace(/[^A-Za-z0-9_]/g,'_').slice(0,40);
  if(typeof XLSX==='undefined'){_csvDl(_rows(ds.recs),safe+'_lineage.csv');return;}
  _xlsxDl([ds],safe+'_lineage.xlsx');
}
function exportAllTabs(){
  const date=new Date().toISOString().slice(0,10);
  if(typeof XLSX==='undefined'){DATASETS.forEach(ds=>{const s=ds.tgt.replace(/[^A-Za-z0-9_]/g,'_').slice(0,40);_csvDl(_rows(ds.recs),s+'_lineage.csv');});return;}
  _xlsxDl(DATASETS,'lineage_multi_'+date+'.xlsx');
}
// ── Column detail panel ────────────────────────────────────────────────────
let _cpData=null,_cpTab='cols';
let _cpEdgeSrc='',_cpEdgeTgt='';
let _cpAllDfs=[];
let _traceMode=false,_traceNodes=new Set(),_traceEdgeKeys=new Set(),_traceCol='';
initTab(0);

function hitEdge(cx,cy){
  let best=null,bestD=10;
  for(let ei=0;ei<EDGES.length;ei++){
    const[s,t]=EDGES[ei];
    const sp=positions[s],tp=positions[t];if(!sp||!tp)continue;
    const[oSlot,iSlot]=edgeSlots[ei]||[0,0];
    const oC=(_outEdges[s]||[]).length,iC=(_inEdges[t]||[]).length;
    const x1=sp.x+NW,y1=sp.y+NH*(oSlot+1)/(oC+1);
    const x2=tp.x,   y2=tp.y+NH*(iSlot+1)/(iC+1);
    const mx=(x1+x2)/2;
    for(let ti=0;ti<=24;ti++){
      const tt=ti/24,u=1-tt;
      const bx=u*u*u*x1+3*u*u*tt*mx+3*u*tt*tt*mx+tt*tt*tt*x2;
      const by=u*u*u*y1+3*u*u*tt*y1+3*u*tt*tt*y2+tt*tt*tt*y2;
      if(Math.hypot(cx-bx,cy-by)<bestD){bestD=Math.hypot(cx-bx,cy-by);best={s,t};}
    }
  }
  return best;
}

function openColPanel(s,t){
  _cpEdgeSrc=s;_cpEdgeTgt=t;
  const ds=DATASETS[currentTab];
  const tShort=t.split('.').pop();
  let colEntry=ds.cols&&(ds.cols[t]||ds.cols[tShort]||ds.cols[t.toUpperCase()]||ds.cols[tShort.toUpperCase()])||null;
  // Fallback: match by job/df from the edge record in RECS
  if((!colEntry||!(colEntry.dataflows||[]).length)&&ds.cols){
    const edgeRec=RECS.find(r=>r.src===s&&r.tgt===t)||RECS.find(r=>r.tgt===t);
    if(edgeRec&&edgeRec.df){
      for(const[,entry] of Object.entries(ds.cols)){
        if((entry.dataflows||[]).some(d=>d.df===edgeRec.df)){colEntry=entry;break;}
      }
    }
  }
  document.getElementById('cpSrc').textContent=shortLabel(s);
  document.getElementById('cpSrc').title=s;
  document.getElementById('cpTgt').textContent=shortLabel(t);
  document.getElementById('cpTgt').title=t;
  const picker=document.getElementById('cpDfPicker');
  const sel=document.getElementById('cpDfSel');
  if(colEntry&&colEntry.dataflows&&colEntry.dataflows.length){
    _cpAllDfs=colEntry.dataflows;
    _cpData=_cpAllDfs[0];
    if(_cpAllDfs.length>1){
      sel.innerHTML=_cpAllDfs.map((d,i)=>'<option value="'+i+'">'+_esc((d.job?d.job+' / ':'')+d.df)+'</option>').join('');
      picker.style.display='';
    }else{
      picker.style.display='none';
    }
  }else{
    _cpAllDfs=[];_cpData=null;
    picker.style.display='none';
  }
  const nM=_cpData?_cpData.mappings.length:0;
  const nJ=_cpData?[...new Set((_cpData.joins||[]).map(j=>j.cond||j.lc))].length:0;
  const nF=_cpData?[...new Set((_cpData.filters||[]).map(f=>f.cond))].length:0;
  document.getElementById('cpCntCols').textContent=nM;
  document.getElementById('cpCntJoins').textContent=nJ;
  document.getElementById('cpCntFilters').textContent=nF;
  document.getElementById('cpDf').textContent=_cpData?(_cpData.df||''):'';
  _cpTab='cols';_updateCpTabs();renderCpBody();
  document.getElementById('colPanel').classList.add('open');
}
function closeColPanel(){
  document.getElementById('colPanel').classList.remove('open');_cpData=null;_cpAllDfs=[];
  document.getElementById('cpDfPicker').style.display='none';
  if(_traceMode){_traceMode=false;_traceNodes=new Set();_traceEdgeKeys=new Set();draw();}
}
let _sqlPanelText='';
function openSqlPanel(id){
  const ds=DATASETS[currentTab]||{};
  const raw=id.startsWith('SQL:')?id.slice(4):id;
  const tfName=raw.includes('|')?raw.split('|').pop():raw;
  const dfName=raw.includes('|')?raw.split('|')[0]:'';
  // Resolve SQL text: try compound key first, then simple key fallback
  _sqlPanelText='';
  if(ds.sql_queries){
    _sqlPanelText=ds.sql_queries[id]||ds.sql_queries['SQL:'+tfName]||'';
  }
  // Source physical tables
  const srcTables=(typeof SQL_MEMBERS!=='undefined'&&SQL_MEMBERS[id])||[];
  // Context labels
  const jobs=[...(nodeJobs&&nodeJobs[id]||[])];
  const ctxParts=[];
  if(dfName)ctxParts.push(dfName);
  if(jobs.length)ctxParts.push(jobs.join(', '));
  document.getElementById('sqlPanelName').textContent=tfName;
  document.getElementById('sqlPanelCtx').textContent=ctxParts.join(' · ');
  // Source tables section
  const srcEl=document.getElementById('sqlSrcSection');
  if(srcTables.length){
    srcEl.style.display='';
    srcEl.innerHTML='<div style="color:#E65100;font-size:11px;font-weight:600;margin-bottom:6px">&#9632; Physical Source Tables ('+srcTables.length+')</div>'
      +srcTables.map(t=>'<span style="display:inline-block;margin:2px 4px 2px 0;padding:3px 9px;background:#fff;border:1px solid #FFAB91;color:#BF360C;border-radius:3px;font-size:11px;font-family:Consolas,monospace">'+_esc(t.split('.').pop())+'</span>').join('');
  }else{
    srcEl.style.display='none';
  }
  // SQL body
  const body=document.getElementById('sqlBody');
  if(_sqlPanelText){
    body.innerHTML='<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">'
      +'<span style="color:#888780;font-size:11px;font-weight:500">SQL Query</span>'
      +'<button id="sqlCopyBtn" onclick="if(_sqlPanelText){navigator.clipboard.writeText(_sqlPanelText).then(()=>{const b=document.getElementById(\'sqlCopyBtn\');b.textContent=\'Copied ✓\';b.style.background=\'#388E3C\';setTimeout(()=>{b.textContent=\'Copy SQL\';b.style.background=\'#D84315\'},2000)})}" '
      +'style="padding:4px 12px;font-size:11px;font-weight:600;background:#D84315;color:#fff;border:none;border-radius:4px;cursor:pointer">Copy SQL</button>'
      +'</div>'
      +'<pre style="margin:0;font-size:11.5px;line-height:1.65;white-space:pre-wrap;word-break:break-word;background:#FFF8F0;padding:14px;border-radius:6px;border:1px solid #FFE0B2;color:#3E0A00;font-family:Consolas,Monaco,monospace">'+_esc(_sqlPanelText)+'</pre>';
  }else{
    body.innerHTML='<div style="color:#888780;font-size:12px;line-height:1.7;padding:20px 0">'
      +'SQL query text was not captured from the XML.<br>'
      +'<br><b style="color:#3E2723">'+_esc(tfName)+'</b> was identified as a SQL Transform node but the query text could not be extracted.'
      +'</div>';
  }
  closeColPanel();
  document.getElementById('sqlPanel').classList.add('open');
}
function closeSqlPanel(){
  document.getElementById('sqlPanel').classList.remove('open');
  _sqlPanelText='';
}
function selectColDf(idx){
  if(!_cpAllDfs[+idx])return;
  _cpData=_cpAllDfs[+idx];
  const nM=_cpData.mappings.length;
  const nJ=[...new Set((_cpData.joins||[]).map(j=>j.cond||j.lc))].length;
  const nF=[...new Set((_cpData.filters||[]).map(f=>f.cond))].length;
  document.getElementById('cpCntCols').textContent=nM;
  document.getElementById('cpCntJoins').textContent=nJ;
  document.getElementById('cpCntFilters').textContent=nF;
  document.getElementById('cpDf').textContent=_cpData.df||'';
  _cpTab='cols';_updateCpTabs();renderCpBody();
}
function showCpTab(tab){_cpTab=tab;_updateCpTabs();renderCpBody();}
function _updateCpTabs(){
  ['cols','joins','filters'].forEach(t=>{
    const id='cpTab'+t.charAt(0).toUpperCase()+t.slice(1);
    document.getElementById(id).classList.toggle('active',t===_cpTab);
  });
}
function renderCpBody(){
  const body=document.getElementById('cpBody');
  if(!_cpData){
    const ds=DATASETS[currentTab];
    const opts=[];
    const tU=(_cpEdgeTgt||'').toUpperCase();
    const tS=(_cpEdgeTgt||'').split('.').pop().toUpperCase();
    // Step 1: only show entries for the clicked target table
    for(const[tbl,entry] of Object.entries(ds&&ds.cols||{})){
      const kU=tbl.toUpperCase(),kS=tbl.split('.').pop().toUpperCase();
      if(kU!==tU&&kS!==tS)continue;
      for(const df of(entry.dataflows||[])){
        opts.push({tbl,df:df.df,job:df.job,entry:df});
      }
    }
    // Step 2: if still empty, match by dataflow name from the RECS edge record
    if(!opts.length){
      const er=RECS.find(r=>r.src===_cpEdgeSrc&&r.tgt===_cpEdgeTgt)||RECS.find(r=>r.tgt===_cpEdgeTgt);
      if(er&&er.df){
        for(const[tbl,entry] of Object.entries(ds&&ds.cols||{})){
          for(const df of(entry.dataflows||[])){
            if(df.df===er.df)opts.push({tbl,df:df.df,job:df.job,entry:df});
          }
        }
      }
    }
    if(!opts.length){
      const lbl=(_cpEdgeTgt||'this table').split('.').pop();
      body.innerHTML='<div class="cp-empty">No column data available for <strong>'+_esc(lbl)+'</strong>.<br>Provide the raw dump XML files in the input folder to extract column-level mappings.</div>';
      return;
    }
    window._cpFallbackOpts=opts;
    body.innerHTML='<div class="cp-fallback">'
      +'<p>No direct column mapping for this edge.<br>Select a Job / Dataflow to view column mappings:</p>'
      +'<select class="cp-fallback-sel" onchange="cpLoadFallback(this.value)">'
      +'<option value="">— select job / dataflow —</option>'
      +opts.map((o,i)=>'<option value="'+i+'">'
        +(o.job?_esc(o.job)+' / ':'')+_esc(o.df)+'</option>').join('')
      +'</select></div>';
    return;
  }
  if(_cpTab==='cols'){
    body.innerHTML='<div class="cp-search-wrap"><input id="cpColSearch" class="cp-search" type="text" placeholder="Filter columns…" oninput="_filterCols()"/></div>'+_renderMappings(_cpData.mappings||[]);
  }else if(_cpTab==='joins'){
    body.innerHTML=_renderJoins(_cpData.joins||[]);
  }else{
    body.innerHTML=_renderFilters(_cpData.filters||[]);
  }
  body.querySelectorAll('.cp-row.has-f').forEach(row=>{
    row.onclick=()=>{const f=row.nextElementSibling;if(f&&f.classList.contains('cp-fx'))f.classList.toggle('open');};
  });
}
function cpLoadFallback(idx){
  const opt=window._cpFallbackOpts&&window._cpFallbackOpts[+idx];
  if(!opt)return;
  _cpData=opt.entry;
  const nM=_cpData.mappings.length;
  const nJ=[...new Set((_cpData.joins||[]).map(j=>j.cond||j.lc))].length;
  const nF=[...new Set((_cpData.filters||[]).map(f=>f.cond))].length;
  document.getElementById('cpCntCols').textContent=nM;
  document.getElementById('cpCntJoins').textContent=nJ;
  document.getElementById('cpCntFilters').textContent=nF;
  document.getElementById('cpDf').textContent=_cpData.df||'';
  _cpTab='cols';_updateCpTabs();renderCpBody();
}
function _filterCols(){
  const q=(document.getElementById('cpColSearch')?.value||'').trim().toLowerCase();
  document.querySelectorAll('#cpBody .cp-row').forEach(row=>{
    const show=!q||row.textContent.toLowerCase().includes(q);
    row.style.display=show?'':'none';
    const fx=row.nextElementSibling;
    if(fx&&fx.classList.contains('cp-fx'))fx.style.display=show?'':'none';
  });
}
function _esc(s){return(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function _renderMappings(mappings){
  if(!mappings.length)return'<div class="cp-empty">No column mappings found for this edge.</div>';
  const byDf={};
  for(const m of mappings){const k=m.df||'(no dataflow)';(byDf[k]||(byDf[k]=[])).push(m);}
  let html='<div class="cp-cols-hdr"><span></span><span style="color:#1D9E75">&#9632; Source</span><span></span><span style="color:#378ADD">&#9632; Target</span></div>';
  for(const[df,rows]of Object.entries(byDf)){
    html+='<div class="cp-df-hdr">'+_esc(df)+'</div>';
    // Group rows by target column to collapse duplicate src→same_tc rows (e.g. SQL JOIN expansion)
    const byTc={};const tcOrder=[];
    for(const m of rows){const k=m.tc||'';if(!byTc[k]){byTc[k]=[];tcOrder.push(k);}byTc[k].push(m);}
    for(const tcKey of tcOrder){
      const tcRows=byTc[tcKey];
      const sorted=[...tcRows].sort((a,b)=>a.d===b.d?0:a.d?-1:1);
      const traceBtn=tcKey?'<button class="cp-trace-btn" data-tbl="'+_esc(_cpEdgeTgt)+'" data-col="'+_esc(tcKey)+'" onclick="event.stopPropagation();traceField(this.dataset.tbl,this.dataset.col)" title="Trace field evolution upstream">&#8679;</button>':'';
      const grpId='tcg_'+Math.random().toString(36).slice(2);
      if(sorted.length>1){
        const tcSpan=tcKey?'<span class="cp-col cp-col-tgt" title="'+_esc(tcKey)+'">'+_esc(tcKey)+'</span>':'<span class="cp-col cp-col-tgt dim">—</span>';
        html+='<div class="cp-tc-grp-hdr" onclick="var b=document.getElementById(\''+grpId+'\');b.classList.toggle(\'open\');this.querySelector(\'.cp-tc-grp-arr\').textContent=b.classList.contains(\'open\')?\'▼\':\'▶\'">'
          +'<span class="cp-tc-grp-arr" style="font-size:9px;color:#888780">▶</span>'
          +'<div style="display:flex;align-items:center;gap:4px">'+tcSpan+traceBtn+'</div>'
          +'<span class="cp-tc-grp-cnt">×'+sorted.length+'</span>'
          +'</div>'
          +'<div class="cp-tc-grp-body" id="'+grpId+'">';
      }
      for(const m of sorted){
        const isDirect=!!m.d,isConst=!m.sc||(m.ft&&m.ft.toUpperCase().includes('CONST'));
        const hasF=!isDirect&&!!m.f;
        const _ft=(m.ft||'').toUpperCase();
        let bs,bl;
        if(_ft==='TABLE_COMPARISON'){bs='background:#EDE9FE;color:#5B21B6';bl='Ck';}
        else if(_ft==='SQL_CUSTOM'){bs='background:#CFFAFE;color:#0E7490';bl='SQ';}
        else if(_ft==='REVERSE_PIVOT'){bs='background:#FFEDD5;color:#C2410C';bl='rP';}
        else if(_ft==='PIVOT'){bs='background:#FEF3C7;color:#B45309';bl='Pv';}
        else if(_ft==='SCD_TYPE2'||_ft==='SCD_TYPE2_METADATA'){bs='background:#FCE7F3;color:#9D174D';bl='H2';}
        else if(_ft==='SURROGATE_KEY'){bs='background:#DBEAFE;color:#1E40AF';bl='Kg';}
        else if(_ft==='ROW_ROUTING'){bs='background:#FEF9C3;color:#92400E';bl='Ca';}
        else if(isDirect){bs='background:#E1F5EE;color:#0F6E56';bl='↔';}
        else if(isConst){bs='background:#E3F2FD;color:#1565C0';bl='C';}
        else{bs='background:#FFF8E1;color:#F57F17';bl='ƒ';}
        const sc=m.sc?'<span class="cp-col cp-col-src" title="'+_esc(m.sc)+'">'+_esc(m.sc)+'</span>'
                     :'<span class="cp-col cp-col-src dim">—</span>';
        // In a group (sorted.length>1), dim the repeated tc cell; show it only in the header
        const tcCell=sorted.length>1?'<span class="cp-col cp-col-tgt dim" style="opacity:0.45;font-size:10px">↑</span>':
          (m.tc?'<span class="cp-col cp-col-tgt" title="'+_esc(m.tc)+'">'+_esc(m.tc)+'</span>':'<span class="cp-col cp-col-tgt dim">—</span>');
        const hint=hasF?'<span style="color:#378ADD;font-size:10px;margin-left:2px">▸</span>':'';
        const rowTraceBtn=sorted.length>1?'':traceBtn;
        html+='<div class="cp-row'+(hasF?' has-f':'')+'">'
          +'<div class="cp-badge" style="'+bs+'">'+bl+'</div>'
          +sc
          +'<div class="cp-arr2">&#8594;</div>'
          +'<div style="display:flex;align-items:center;min-width:0">'+tcCell+hint+rowTraceBtn+'</div>'
          +'</div>';
        if(hasF)html+='<div class="cp-fx">'+_esc(m.f)+'</div>';
      }
      if(sorted.length>1)html+='</div>';
    }
  }
  return html;
}
function _renderJoins(joins){
  if(!joins||!joins.length)return'<div class="cp-empty">No join conditions found.</div>';
  let html='';const seen=new Set();
  for(const j of joins){
    const cond=j.cond||(j.lc&&j.rc?j.lo+'.'+j.lc+' = '+j.ro+'.'+j.rc:'');
    if(!cond||seen.has(cond))continue;seen.add(cond);
    html+='<div class="cp-join">'
      +'<div class="cp-jtype">'+(j.jt||'JOIN')+'</div>'
      +(j.df?'<div class="cp-jdf">'+_esc(j.df)+'</div>':'')
      +'<div class="cp-jcond">'+_esc(cond)+'</div>'
      +'</div>';
  }
  return html||'<div class="cp-empty">No join conditions found.</div>';
}
function _renderFilters(filters){
  if(!filters||!filters.length)return'<div class="cp-empty">No filter conditions found.</div>';
  let html='';const seen=new Set();
  for(const f of filters){
    if(!f.cond||seen.has(f.cond))continue;seen.add(f.cond);
    html+='<div class="cp-flt">'
      +(f.df?'<div class="cp-flabel">'+_esc(f.df)+'</div>':'')
      +'<div class="cp-fcond">'+_esc(f.cond)+'</div>'
      +'</div>';
  }
  return html||'<div class="cp-empty">No filter conditions found.</div>';
}

// ── Field Evolution / Column Trace ─────────────────────────────────────────
function _clearTrace(){
  _traceMode=false;_traceNodes=new Set();_traceEdgeKeys=new Set();_traceCol='';
}
function buildFieldTrace(table,col,depth,visited){
  depth=depth==null?0:depth;visited=visited||new Set();
  if(depth>10)return{table,col,formula:'[max depth]',df:'',children:[]};
  const vk=table+'|'+col.toUpperCase();
  if(visited.has(vk))return null;
  visited=new Set(visited);visited.add(vk);
  const ds=DATASETS[currentTab];
  const node={table,col,formula:'',df:'',children:[]};
  const tblU=table.toUpperCase(),tblShort=table.split('.').pop().toUpperCase();
  for(const[tblKey,tblData]of Object.entries(ds.cols||{})){
    const tkU=tblKey.toUpperCase(),tkShort=tblKey.split('.').pop().toUpperCase();
    if(tkU!==tblU&&tkShort!==tblShort)continue;
    for(const dfEntry of(tblData.dataflows||[])){
      for(const m of(dfEntry.mappings||[])){
        if(!m.tc||m.tc.toUpperCase()!==col.toUpperCase())continue;
        if(!m.sc)continue;
        if(!node.formula&&m.f)node.formula=m.f;
        if(!node.df&&m.df)node.df=m.df;
        const child=buildFieldTrace(m.src_obj||tblKey,m.sc,depth+1,visited);
        if(child)node.children.push(child);
      }
    }
  }
  return node;
}
function _collectTraceSet(node){
  const nodes=new Set(),edges=new Set();
  function walk(n){
    nodes.add(n.table);
    for(const c of n.children){
      edges.add(c.table+'→'+n.table);
      walk(c);
    }
  }
  walk(node);
  return{nodes,edges};
}
function traceField(table,col){
  _traceCol=col;
  const tree=buildFieldTrace(table,col);
  if(!tree)return;
  const{nodes,edges}=_collectTraceSet(tree);
  _traceMode=true;_traceNodes=nodes;_traceEdgeKeys=edges;
  document.getElementById('cpTraceLbl').textContent='Field: '+col;
  document.getElementById('cpNormalTabs').style.display='none';
  document.getElementById('cpTraceHdr').classList.add('open');
  document.getElementById('cpBody').innerHTML=_renderTraceTree(tree,0);
  draw();
}
function exitTraceMode(){
  _clearTrace();
  document.getElementById('cpNormalTabs').style.display='';
  document.getElementById('cpTraceHdr').classList.remove('open');
  _cpTab='cols';_updateCpTabs();renderCpBody();
  draw();
}
function _renderTraceTree(node,depth){
  const isTerminal=node.children.length===0;
  const icon=depth===0?'&#127919;':isTerminal?'&#127807;':'&#8599;';
  const tblShort=node.table.split('.').pop();
  let h='<div class="cp-te-node"'+(depth>0?' style="margin-left:'+(depth*14)+'px"':'')+'>'+
    '<div class="cp-te-row">'+
      '<span class="cp-te-icon">'+icon+'</span>'+
      '<span class="cp-te-tbl" title="'+_esc(node.table)+'">'+_esc(tblShort)+'</span>'+
      '<span class="cp-te-col'+(isTerminal?' cp-te-terminal':'')+'">'+_esc(node.col)+'</span>'+
      (node.df?'<span class="cp-te-df">'+_esc(node.df)+'</span>':'')+
    '</div>'+
    (node.formula&&depth>0?'<div class="cp-te-fx">'+_esc(node.formula)+'</div>':'')+
  '</div>';
  for(const child of node.children)h+=_renderTraceTree(child,depth+1);
  return h;
}
</script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
</body>
</html>"""


class DSXMLLauncher:
    def __init__(self, root):
        self.root = root
        self.root.title("DS XML Lineage Extractor")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # Variables
        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.is_running = False
        self.process_queue = queue.Queue()
        self.max_log_lines = 2000

        # Disk-based extraction caching (no memory bloat)
        self.extraction_csv_path = None  # Path to extracted CSV written to disk
        self.table_lineage_csv_path = None  # Path to table lineage CSV

        # Targeted lineage search state
        self.targeted_is_running = False
        self.targeted_queue = queue.Queue()
        self.targeted_results = []             # list of result dicts (one per target table)
        self.targeted_xml_var = tk.StringVar()
        self.targeted_table_var = tk.StringVar()
        self.targeted_output_var = tk.StringVar()
        self.targeted_hops_var = tk.IntVar(value=10)
        self.graph_btn = None   # assigned in _create_targeted_tab; enabled after run

        # Raw dump path — shared by Tab 2 (Targeted) and Tab 3 (Job)
        self.raw_dump_path_var = tk.StringVar()

        # Job-level search state (Tab 3)
        self.job_is_running = False
        self.job_queue = queue.Queue()
        self.job_results = []
        self.job_xml_var = tk.StringVar()
        self.job_name_var = tk.StringVar()
        self.job_output_var = tk.StringVar()
        self.job_hops_var = tk.IntVar(value=10)
        self.job_graph_btn = None  # assigned in _create_job_tab; enabled after run

        # Column Lineage tab (Tab 4) state
        self.col_tree        = None  # assigned in _create_column_tab
        self.col_summary_var = None
        self.col_filter_var  = None
        self.col_export_btn  = None
        self._col_all_rows: list = []  # cached (tgt, row) pairs for filter

        # Shared input folder — auto-detects BODS_FULL_PROD.xml + both raw dumps
        self.input_folder_var = tk.StringVar()
        self.resolved_raw_dump_paths: list = []   # populated by _resolve_input_folder

        # Recent search history and per-run timing
        self.targeted_recent_tables: list = []
        self.job_recent_tables: list = []
        self._targeted_start_time: float = 0.0
        self._job_start_time: float = 0.0

        # Load default paths from config if available
        self.load_default_paths()

        # Create UI
        self.create_widgets()

        # Setup logging to redirect to GUI
        self.setup_logging()

    def load_default_paths(self):
        """Pick sensible default paths.

        Prefer the paths in config.py, but only if they actually exist on THIS
        machine — config ships with the developer's personal paths, which are
        meaningless on an end user's PC (especially when running as a packaged
        .exe). Otherwise fall back to the user's Documents folder so Browse
        dialogs open somewhere familiar.
        """
        home = os.path.expanduser("~")
        documents = os.path.join(home, "Documents")
        base = documents if os.path.isdir(documents) else home
        default_output = os.path.join(base, "DS_Lineage_Output")

        cfg_input = cfg_output = None
        try:
            sys.path.append(os.getcwd())
            import config
            if hasattr(config, 'INPUT'):
                cfg_input = config.INPUT.get('path')
            if hasattr(config, 'OUTPUT'):
                cfg_output = config.OUTPUT.get('output_dir')
        except Exception:
            pass

        self.input_path.set(cfg_input if (cfg_input and os.path.exists(cfg_input)) else base)
        self.output_path.set(cfg_output if (cfg_output and os.path.isdir(cfg_output)) else default_output)
        self.targeted_output_var.set(cfg_output if (cfg_output and os.path.isdir(cfg_output)) else default_output)
        self.job_output_var.set(cfg_output if (cfg_output and os.path.isdir(cfg_output)) else default_output)

        # Override defaults with saved session state (last-used paths win over config defaults)
        session = self._load_session_state()
        if session.get("input_folder") and os.path.isdir(session["input_folder"]):
            self.input_folder_var.set(session["input_folder"])
            self._resolve_input_folder(session["input_folder"])
        if session.get("targeted_output"):
            self.targeted_output_var.set(session["targeted_output"])
        if "targeted_hops" in session:
            try:
                self.targeted_hops_var.set(int(session["targeted_hops"]))
            except Exception:
                pass
        self.targeted_recent_tables = session.get("targeted_recent", [])
        if session.get("job_output"):
            self.job_output_var.set(session["job_output"])
        if "job_hops" in session:
            try:
                self.job_hops_var.set(int(session["job_hops"]))
            except Exception:
                pass
        self.job_recent_tables = session.get("job_recent", [])
        if session.get("main_output"):
            self.output_path.set(session["main_output"])
        if session.get("geometry"):
            try:
                self.root.geometry(session["geometry"])
            except Exception:
                pass

    # ── Session state persistence ─────────────────────────────────────────

    def _session_state_path(self) -> str:
        return os.path.join(os.path.expanduser("~"), ".ds_lineage_session.json")

    def _load_session_state(self) -> dict:
        try:
            with open(self._session_state_path(), encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_session_state(self):
        try:
            state = {
                "input_folder":    self.input_folder_var.get(),
                "targeted_output": self.targeted_output_var.get(),
                "targeted_hops":   self.targeted_hops_var.get(),
                "targeted_recent": self.targeted_recent_tables,
                "job_output":      self.job_output_var.get(),
                "job_hops":        self.job_hops_var.get(),
                "job_recent":      self.job_recent_tables,
                "main_output":     self.output_path.get(),
                "geometry":        self.root.geometry(),
            }
            with open(self._session_state_path(), "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2)
        except Exception:
            pass

    def _on_close(self):
        self._save_session_state()
        self.root.destroy()

    # ── Input folder auto-detection ───────────────────────────────────────

    _MAIN_XML   = "BODS_FULL_PROD.xml"
    _RAW_DF     = "raw_DATAFLOW_FULL_PROD.xml"
    _RAW_ABAP   = "raw_ABAP_DATAFLOW_FULL_PROD.xml"

    def _browse_input_folder(self):
        current = self.input_folder_var.get()
        folder = filedialog.askdirectory(
            title="Select Input Folder (must contain BODS_FULL_PROD.xml)",
            initialdir=current if os.path.isdir(current) else os.path.expanduser("~/Documents"),
        )
        if folder:
            self.input_folder_var.set(folder)
            self._resolve_input_folder(folder)

    def _resolve_input_folder(self, folder: str):
        """Detect the three expected XML files in folder and wire up all path variables."""
        main_xml  = os.path.join(folder, self._MAIN_XML)
        raw_df    = os.path.join(folder, self._RAW_DF)
        raw_abap  = os.path.join(folder, self._RAW_ABAP)

        if os.path.isfile(main_xml):
            self.input_path.set(folder)          # Tab 1 Full Extraction uses the folder
            self.targeted_xml_var.set(main_xml)  # Tab 2 Targeted Lineage
            self.job_xml_var.set(main_xml)        # Tab 3 Job Level

        self.resolved_raw_dump_paths = [p for p in (raw_df, raw_abap) if os.path.isfile(p)]

        if hasattr(self, "folder_status_var"):
            parts = []
            for label, path in [
                (self._MAIN_XML,  main_xml),
                (self._RAW_DF,    raw_df),
                (self._RAW_ABAP,  raw_abap),
            ]:
                icon = "✓" if os.path.isfile(path) else "✗"
                # Shorten label for display
                short = label.replace("raw_DATAFLOW_FULL_PROD", "raw_DATAFLOW") \
                             .replace("raw_ABAP_DATAFLOW_FULL_PROD", "raw_ABAP") \
                             .replace("BODS_FULL_PROD", "BODS_FULL_PROD")
                parts.append(f"{icon} {short}")
            self.folder_status_var.set("   ".join(parts))

    def create_widgets(self):
        """Create the main UI widgets"""
        # Root grid — row 0 = shared folder picker, row 1 = notebook
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=0)
        self.root.rowconfigure(1, weight=1)

        # ── Shared Input Folder picker (above all tabs) ───────────────────
        folder_frame = ttk.LabelFrame(self.root, text=" Input Folder ", padding="6 4")
        folder_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=8, pady=(6, 0))
        folder_frame.columnconfigure(1, weight=1)

        ttk.Label(folder_frame, text="Folder:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(folder_frame, textvariable=self.input_folder_var, width=70).grid(
            row=0, column=1, sticky=(tk.W, tk.E), padx=(8, 6))
        ttk.Button(folder_frame, text="Browse…",
                   command=self._browse_input_folder).grid(row=0, column=2)

        self.folder_status_var = tk.StringVar(value="No folder selected — browse to the folder containing the three XML files")
        ttk.Label(folder_frame, textvariable=self.folder_status_var,
                  font=("Consolas", 8), foreground="#555555").grid(
            row=1, column=0, columnspan=3, sticky=tk.W, pady=(2, 0))

        # Re-run folder resolution now that folder_status_var exists
        if self.input_folder_var.get():
            self._resolve_input_folder(self.input_folder_var.get())

        # Tabbed notebook — Tab 1 = Full Extraction, Tab 2 = Targeted Lineage
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        tab1 = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab1, text="  Full Extraction  ")

        main_frame = tab1
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(7, weight=1)

        # Title
        title_label = ttk.Label(main_frame, text="SAP Data Services XML Lineage Extractor",
                               font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # Input directory selection
        ttk.Label(main_frame, text="Input Path:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.input_entry = ttk.Entry(main_frame, textvariable=self.input_path, width=60)
        self.input_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=5)
        ttk.Button(main_frame, text="Browse...", command=self.browse_input).grid(row=1, column=2, padx=(5, 0), pady=5)

        # Output directory selection
        ttk.Label(main_frame, text="Output Path:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_entry = ttk.Entry(main_frame, textvariable=self.output_path, width=60)
        self.output_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=5)
        ttk.Button(main_frame, text="Browse...", command=self.browse_output).grid(row=2, column=2, padx=(5, 0), pady=5)

        # Batch export mode checkbox
        self.batch_mode_var = tk.BooleanVar(value=False)
        batch_check = ttk.Checkbutton(main_frame, text="Batch Mode (unwrap + progressive export)", 
                                     variable=self.batch_mode_var)
        batch_check.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=5)

        # Control buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=3, pady=10)

        self.run_button = ttk.Button(button_frame, text="Run Extraction", command=self.run_extraction)
        self.run_button.pack(side=tk.LEFT, padx=(0, 10))

        self.stop_button = ttk.Button(button_frame, text="Stop", command=self.stop_extraction, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Button(button_frame, text="Clear Log", command=self.clear_log).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Save Log", command=self.save_log).pack(side=tk.LEFT, padx=(0, 10))

        # Export buttons
        ttk.Label(button_frame, text="Export:").pack(side=tk.LEFT, padx=(20, 5))
        ttk.Button(button_frame, text="📄 Column-Level", command=self._export_column_level).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="📊 Table-Level", command=self._export_table_level).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="📦 Full Lineage", command=self._export_full_lineage).pack(side=tk.LEFT)

        # Progress bar with 3-decimal display
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(1, weight=1)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5))

        self.progress_label = ttk.Label(progress_frame, text="0.000%", width=10, anchor=tk.E)
        self.progress_label.grid(row=0, column=2, padx=(0, 0))

        # Log display
        ttk.Label(main_frame, text="Log Output:").grid(row=6, column=0, sticky=tk.W, pady=(10, 5))
        self.log_text = scrolledtext.ScrolledText(main_frame, height=20, wrap=tk.WORD,
                                                 font=("Consolas", 9))
        self.log_text.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E))

        # Tab 2: Targeted Lineage Search
        self._create_targeted_tab()

        # Tab 3: Job Level Analysis
        self._create_job_tab()

        # Tab 4: Column Lineage — field-level mapping viewer
        self._create_column_tab()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def browse_input(self):
        """Browse for input directory or file"""
        current_path = self.input_path.get()
        if os.path.isfile(current_path):
            initial_dir = os.path.dirname(current_path)
            file_path = filedialog.askopenfilename(
                title="Select DS XML File",
                initialdir=initial_dir,
                filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
            )
            if file_path:
                self.input_path.set(file_path)
        else:
            dir_path = filedialog.askdirectory(
                title="Select Input Directory",
                initialdir=current_path if os.path.exists(current_path) else os.getcwd()
            )
            if dir_path:
                self.input_path.set(dir_path)

    def browse_output(self):
        """Browse for output directory"""
        current_path = self.output_path.get()
        dir_path = filedialog.askdirectory(
            title="Select Output Directory",
            initialdir=current_path if os.path.exists(current_path) else os.getcwd()
        )
        if dir_path:
            self.output_path.set(dir_path)

    def setup_logging(self):
        """Setup logging to redirect output to the GUI"""
        # Create a custom handler that writes to the GUI
        class GUITextHandler(logging.Handler):
            def __init__(self, text_widget):
                super().__init__()
                self.text_widget = text_widget

            def emit(self, record):
                msg = self.format(record)
                self.text_widget.after(0, self._append_text, msg + '\n')

            def _append_text(self, msg):
                self.text_widget.insert(tk.END, msg)
                self.text_widget.see(tk.END)

        # Remove existing handlers and install GUI logging only
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)

        gui_handler = GUITextHandler(self.log_text)
        gui_handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)-8s | %(message)s',
                                                 datefmt='%H:%M:%S'))
        logging.root.addHandler(gui_handler)
        logging.root.setLevel(logging.INFO)

        # Prevent engine-specific logger from duplicating messages through root handler
        logging.getLogger("ds-xml-lineage").propagate = False

    def _append_log_message(self, message: str):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        if int(self.log_text.index('end-1c').split('.')[0]) > self.max_log_lines:
            self.log_text.delete('1.0', '201.0')

    def _enqueue_engine_message(self, message, level="info"):
        if level == 'progress':
            # Keep the status text and update the progress bar numeric value with 3 decimals
            self.process_queue.put(('status', message))
            try:
                percent = float(re.search(r"(\d+(?:\.\d+)?)%", message).group(1))
                self.process_queue.put(('progress', percent))
            except Exception:
                pass
            return

        if level in ('status', 'info', 'warning', 'error', 'success'):
            self.process_queue.put((level, message))
        else:
            self.process_queue.put(('info', message))

    def run_extraction(self):
        """Run the DS XML extraction directly in a worker thread (no subprocess)"""
        if self.is_running:
            return

        input_path = self.input_path.get().strip()
        output_path = self.output_path.get().strip()

        if not input_path:
            messagebox.showerror("Error", "Please select an input path")
            return

        if not output_path:
            messagebox.showerror("Error", "Please select an output path")
            return

        # Create output directory if it doesn't exist
        os.makedirs(output_path, exist_ok=True)

        # Clear previous log
        self.log_text.delete(1.0, tk.END)
        
        # Add initial status message to log
        self.log_text.insert(tk.END, f"="*80 + "\n")
        self.log_text.insert(tk.END, f"[START] DS XML Lineage Extraction\n")
        self.log_text.insert(tk.END, f"Input: {input_path}\n")
        self.log_text.insert(tk.END, f"Output: {output_path}\n")
        self.log_text.insert(tk.END, f"="*80 + "\n\n")
        self.log_text.insert(tk.END, f"[INFO] Starting extraction in worker thread...\n")
        self.log_text.insert(tk.END, f"[INFO] Loading ds_xml_engine (this may take 10-20 seconds)...\n\n")
        self.log_text.see(tk.END)
        self.root.update()  # Force UI update

        # Update UI state
        self.is_running = True
        self.run_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_var.set("Loading extraction engine (worker thread)...")
        self.progress_var.set(5)
        self.progress_label.config(text="5.000%")

        # Direct threaded approach - no subprocess
        extraction_thread = threading.Thread(target=self._run_extraction_thread,
                                           args=(input_path, output_path))
        extraction_thread.daemon = True
        extraction_thread.start()
        self.root.after(100, self._check_extraction_status)

    def _run_extraction_thread(self, input_path, output_path):
        """Run the extraction process (direct in-process call)"""
        try:
            self.process_queue.put(('status', 'Loading ds_xml_engine (heavy imports)...'))

            # Import the DS engine
            import ds_xml_engine
            self.process_queue.put(('info', 'Successfully imported ds_xml_engine'))
            self.process_queue.put(('progress', 10.0))

            # Check if batch mode is enabled
            use_batch_mode = self.batch_mode_var.get()
            
            if use_batch_mode:
                self.process_queue.put(('status', 'Starting BATCH MODE extraction with progressive export...'))
                self.process_queue.put(('info', 'Batch Mode: unwrapping elements and streaming to CSV'))
                self.process_queue.put(('progress', 15.0))
                
                # Get all XML files
                if os.path.isfile(input_path) and input_path.lower().endswith('.xml'):
                    xml_files = [input_path]
                elif os.path.isdir(input_path):
                    xml_files = [os.path.join(root, file)
                                for root, _, files in os.walk(input_path)
                                for file in files if file.lower().endswith('.xml')]
                else:
                    self.process_queue.put(('error', f'Invalid input path: {input_path}'))
                    return
                
                if not xml_files:
                    self.process_queue.put(('error', f'No XML files found in: {input_path}'))
                    return
                
                self.process_queue.put(('info', f'Found {len(xml_files)} XML file(s)'))
                
                # Process each file in batch mode
                for idx, xml_file in enumerate(sorted(xml_files), 1):
                    percent = 15 + (idx / len(xml_files)) * 75
                    self.process_queue.put(('progress', percent))
                    self.process_queue.put(('status', f'[{idx}/{len(xml_files)}] Processing {os.path.basename(xml_file)} (batch mode)...'))
                    
                    rows_written, elem_counts, fsize = ds_xml_engine._extract_streaming_batch_export(
                        xml_file, output_path, 
                        status_callback=self._enqueue_engine_message
                    )
                    
                    self.process_queue.put(('info', f'  >> Batch export complete: {rows_written} rows written'))
                
                self.process_queue.put(('progress', 95.0))
                self.process_queue.put(('success', 'Batch extraction complete! CSV files written progressively to disk'))
                self.process_queue.put(('progress', 100.0))
            else:
                # Standard extraction mode
                self.process_queue.put(('status', 'Starting standard extraction...'))
                self.process_queue.put(('progress', 15.0))
                
                rows, lineage_graph, table_lineage_df, processed_files = ds_xml_engine.extract_ds_lineage_from_path(
                    input_path,
                    status_callback=self._enqueue_engine_message
                )
                
                self.process_queue.put(('progress', 60.0))
                self.process_queue.put(('info', 'Extraction phase completed'))

                if not processed_files:
                    self.process_queue.put(('error', f'No XML files found in {input_path}'))
                    return

                if rows and len(rows) > 0:
                    self.process_queue.put(('progress', 70.0))
                    self.process_queue.put(('status', f'Building table-level lineage from {len(rows)} rows...'))
                    
                    table_lineage_df = extract_table_lineage(rows)
                    self.process_queue.put(('info', f"✅ Table lineage built: {len(table_lineage_df)} rows"))
                    
                    self.process_queue.put(('progress', 80.0))
                    self.process_queue.put(('status', 'Exporting results to CSV...'))
                    
                    # Export only to CSV (skip Excel which is slow) - CSV is much faster
                    # For UI use, CSV is preferred since it's quick and easier to share
                    try:
                        import csv

                        col_path = None
                        table_path = None
                        summary_path = None

                        if rows:
                            self.process_queue.put(('status', 'Exporting column lineage to CSV...'))
                            self.process_queue.put(('progress', 82.0))

                            try:
                                import config
                                delimiter = config.SMART_EXPORT.get('csv_delimiter', ',') if hasattr(config, 'SMART_EXPORT') else ','
                            except ImportError:
                                delimiter = ','

                            # Use get_column_mapping_df so the cached CSV has a
                            # stable, readable column order (Source_Object,
                            # Source_Column, Source_Column_Name, Target_Object,
                            # Target_Column, Target_Column_Name, then the rest).
                            # The Full/Column/Table exports all read this cache,
                            # so ordering it here fixes ordering everywhere.
                            df = ds_xml_engine.get_column_mapping_df(rows)
                            if df.empty:
                                df = pd.DataFrame(rows)
                            col_path = os.path.join(output_path, "ds_column_lineage.csv")
                            df.to_csv(col_path, index=False, sep=delimiter, encoding='utf-8')
                            self.process_queue.put(('info', f"✅ Column CSV exported: {len(df)} rows"))

                        self.process_queue.put(('progress', 88.0))
                        self.process_queue.put(('status', 'Writing table lineage CSV...'))

                        if table_lineage_df is not None and not table_lineage_df.empty:
                            table_path = os.path.join(output_path, "ds_table_lineage.csv")
                            table_lineage_df.to_csv(table_path, index=False, sep=delimiter)
                            self.process_queue.put(('info', f"✅ Table CSV exported: {len(table_lineage_df)} rows"))

                        export_paths = {
                            'column_lineage_csv': col_path,
                            'table_lineage_csv': table_path,
                            'summary_csv': summary_path,
                            'excel_workbook': None  # Skip Excel - it's slow for UI
                        }
                    except Exception as export_err:
                        self.process_queue.put(('error', f'Export error: {str(export_err)}'))
                        import traceback
                        for line in traceback.format_exc().split('\n'):
                            if line.strip():
                                self.process_queue.put(('error', line))

                    # Store CSV file paths for disk-based export (no memory bloat)
                    if 'column_lineage_csv' in export_paths and export_paths['column_lineage_csv']:
                        self.extraction_csv_path = export_paths['column_lineage_csv']
                        self.process_queue.put(('info', f"✅ Ready for Excel export (cached CSV)"))

                    self.process_queue.put(('progress', 95.0))
                    self.process_queue.put((
                        'success',
                        f"Extraction complete! {len(processed_files)} XML file(s), {len(rows)} total rows extracted"
                    ))

                    self.process_queue.put(('info', f'Processed {len(processed_files)} XML file(s)'))
                    for key, path in export_paths.items():
                        if path:
                            key_display = key.replace('_', ' ').title()
                            self.process_queue.put(('info', f'{key_display}: {path}'))
                    
                    self.process_queue.put(('progress', 100.0))
                else:
                    self.process_queue.put(('warning', 'No data extracted from any files'))

        except Exception as e:
            import traceback
            err_msg = f'Extraction failed: {str(e)}'
            self.process_queue.put(('error', err_msg))
            tb = traceback.format_exc()
            for line in tb.split('\n'):
                if line.strip():
                    self.process_queue.put(('error', line))
        finally:
            self.process_queue.put(('done', None))

    def _save_results(self, rows, output_path):
        """Save extraction results to files"""
        try:
            import pandas as pd
            import json

            # Convert to DataFrame with the standard source/target-first column
            # order (same ordering as every other export path).
            import ds_xml_engine
            df = ds_xml_engine.get_column_mapping_df(rows)
            if df.empty:
                df = pd.DataFrame(rows)

            # Save as CSV
            csv_path = os.path.join(output_path, "ds_lineage_results.csv")
            df.to_csv(csv_path, index=False)
            self.process_queue.put(('info', f'Results saved to: {csv_path}'))

            # Save as Excel if possible with separate sheets for easier browsing
            try:
                excel_path = os.path.join(output_path, "ds_lineage_results.xlsx")
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='All_Rows', index=False)

                    mapping_df = df[df['Record_Type'] == 'COLUMN_MAPPING']
                    if not mapping_df.empty:
                        mapping_df.to_excel(writer, sheet_name='Column_Mappings', index=False)

                    projection_df = df[df['Record_Type'] == 'PROJECTION']
                    if not projection_df.empty:
                        projection_df.to_excel(writer, sheet_name='Projections', index=False)

                    query_df = df[df['Record_Type'].isin(['JOIN', 'FILTER', 'GROUP_BY', 'ORDER_BY', 'UNNEST'])]
                    if not query_df.empty:
                        query_df.to_excel(writer, sheet_name='Query_Level', index=False)

                self.process_queue.put(('info', f'Excel results saved to: {excel_path}'))
            except ImportError:
                self.process_queue.put(('warning', 'Excel export not available (openpyxl not installed)'))
            except Exception as e:
                self.process_queue.put(('warning', f'Excel export failed: {str(e)}'))

            # Save summary
            proj_count = len([r for r in rows if r.get('Record_Type') == 'PROJECTION'])
            mapped_count = len([r for r in rows if r.get('Record_Type') == 'COLUMN_MAPPING'])
            extraction_coverage_pct = None
            if proj_count:
                try:
                    extraction_coverage_pct = round((mapped_count / proj_count) * 100.0, 2)
                except Exception:
                    extraction_coverage_pct = None

            summary = {
                'total_rows': len(rows),
                'column_mappings': mapped_count,
                'projections': proj_count,
                'expected_output_columns': proj_count,
                'extraction_coverage_pct': extraction_coverage_pct,
                'query_level_rows': len([r for r in rows if r.get('Record_Type') in {'JOIN', 'FILTER', 'GROUP_BY', 'ORDER_BY', 'UNNEST'}]),
                'unique_source_objects': len(set(r.get('Source_Object', '') for r in rows if r.get('Source_Object'))),
                'unique_target_objects': len(set(r.get('Target_Object', '') for r in rows if r.get('Target_Object'))),
                'output_files': [csv_path]
            }

            summary_path = os.path.join(output_path, "extraction_summary.json")
            with open(summary_path, 'w') as f:
                json.dump(summary, f, indent=2)
            self.process_queue.put(('info', f'Summary saved to: {summary_path}'))

        except Exception as e:
            self.process_queue.put(('error', f'Error saving results: {str(e)}'))

    def _export_column_level(self):
        """Export column-level lineage"""
        if self.extraction_csv_path is None:
            messagebox.showwarning("No Data", "Please run extraction first")
            return
        self._do_export('COLUMN_LEVEL')

    def _export_table_level(self):
        """Export table-level lineage"""
        if self.extraction_csv_path is None:
            messagebox.showwarning("No Data", "Please run extraction first")
            return
        self._do_export('TABLE_LEVEL')

    def _export_full_lineage(self):
        """Export full lineage"""
        if self.extraction_csv_path is None:
            messagebox.showwarning("No Data", "Please run extraction first")
            return
        self._do_export('FULL')

    def _do_export(self, lineage_level):
        """Perform the export with specified lineage level by reading from cached CSV"""
        output_path = self.output_path.get().strip()
        if not output_path:
            messagebox.showerror("Error", "Please select an output path")
            return

        os.makedirs(output_path, exist_ok=True)
        
        try:
            import ds_xml_engine
            
            # Read cached extraction from disk
            try:
                import config
                delimiter = config.SMART_EXPORT.get('csv_delimiter', ',') if hasattr(config, 'SMART_EXPORT') else ','
            except ImportError:
                delimiter = ','

            df = pd.read_csv(self.extraction_csv_path, sep=delimiter)
            rows = df.to_dict('records')
            
            # Reconstruct DataFrame after filtering to preserve data types
            if lineage_level == 'COLUMN_LEVEL':
                filtered_rows = ds_xml_engine.filter_column_level_rows(rows)
                filtered_df = pd.DataFrame(filtered_rows) if filtered_rows else pd.DataFrame()
            elif lineage_level == 'TABLE_LEVEL':
                filtered_rows = ds_xml_engine.filter_table_level_rows(rows)
                filtered_df = pd.DataFrame(filtered_rows) if filtered_rows else pd.DataFrame()
            else:  # FULL
                filtered_df = df
            
            if filtered_df.empty:
                messagebox.showwarning("No Data", f"No {lineage_level} lineage data to export")
                return
            
            # Save filtered results to CSV and Excel
            output_csv = os.path.join(output_path, f"ds_lineage_{lineage_level.lower()}.csv")
            filtered_df.to_csv(output_csv, index=False, sep=delimiter)
            
            # Try to save Excel too
            try:
                output_xlsx = os.path.join(output_path, f"ds_lineage_{lineage_level.lower()}.xlsx")
                filtered_df.to_excel(output_xlsx, index=False, engine='openpyxl')
                messagebox.showinfo("Export Success", f"{lineage_level} lineage exported:\n  CSV: {output_csv}\n  XLSX: {output_xlsx}")
            except Exception as e:
                messagebox.showinfo("Export Success", f"{lineage_level} lineage exported to:\n  CSV: {output_csv}\n(Excel export failed: {str(e)})")
            
            logging.info(f"Export {lineage_level}: {output_csv}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export {lineage_level}: {str(e)}")
            logging.error(f"Export {lineage_level} error: {str(e)}", exc_info=True)

    def _check_extraction_status(self):
        """Check for messages from the extraction thread"""
        try:
            while True:
                msg_type, msg = self.process_queue.get_nowait()

                if msg_type == 'status':
                    self.status_var.set(msg)
                    self.log_text.insert(tk.END, f"[STATUS] {msg}\n")
                elif msg_type == 'info':
                    self._append_log_message(f"[INFO] {msg}")
                elif msg_type == 'warning':
                    self._append_log_message(f"[WARN] {msg}")
                elif msg_type == 'error':
                    self._append_log_message(f"[ERROR] {msg}")
                elif msg_type == 'success':
                    self.status_var.set("Extraction completed successfully")
                    self._append_log_message(f"[SUCCESS] {msg}")
                elif msg_type == 'progress':
                    try:
                        percent = float(msg)
                        self.progress_var.set(percent)
                        # Update label with 3 decimal places
                        self.progress_label.config(text=f"{percent:.3f}%")
                    except Exception:
                        pass
                elif msg_type == 'done':
                    self._extraction_finished()
                    return

                # Auto-scroll log to bottom
                self.log_text.see(tk.END)

        except queue.Empty:
            pass

        # Continue checking
        if self.is_running:
            self.root.after(100, self._check_extraction_status)

    def _extraction_finished(self):
        """Handle extraction completion"""
        self.is_running = False
        self.run_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.progress_var.set(100)
        # Cleanup - threaded extraction will finish on its own
        self.proc = None

    def stop_extraction(self):
        """Stop the current extraction process"""
        if self.is_running:
            self.is_running = False
            self.process_queue.put(('warning', 'Extraction stopped by user'))
            self._extraction_finished()


    def clear_log(self):
        """Clear the log display"""
        self.log_text.delete(1.0, tk.END)

    def save_log(self):
        """Save the log output to a file"""
        log_content = self.log_text.get(1.0, tk.END).strip()
        if not log_content:
            messagebox.showinfo("Info", "No log content to save")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save Log File",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("Log files", "*.log"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(log_content)
                messagebox.showinfo("Success", f"Log saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save log: {str(e)}")


    # ======================================================================
    # Targeted Lineage Tab
    # ======================================================================

    def _create_targeted_tab(self):
        """Build the Targeted Lineage Search tab (Tab 2)."""
        tab2 = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab2, text="  Targeted Lineage  ")

        tab2.columnconfigure(1, weight=1)
        tab2.rowconfigure(6, weight=1)

        # ── Row 0: Output directory ──────────────────────────────────────
        ttk.Label(tab2, text="Output Dir:").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.t_output_entry = ttk.Entry(tab2, textvariable=self.targeted_output_var, width=55)
        self.t_output_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=4)
        ttk.Button(tab2, text="Browse…",
                   command=self._browse_targeted_output).grid(row=0, column=2, padx=(5, 0), pady=4, sticky=tk.W)

        # ── Row 1: Target table(s) ───────────────────────────────────────
        ttk.Label(tab2, text="Target Table(s):").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.t_table_entry = ttk.Combobox(tab2, textvariable=self.targeted_table_var, width=55)
        self.t_table_entry['values'] = self.targeted_recent_tables
        self.t_table_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=4)
        ttk.Label(tab2, text="comma-separated for multi-tab graph", foreground="grey").grid(
            row=1, column=2, sticky=tk.W, padx=(5, 0))

        # ── Row 2: Options + controls ────────────────────────────────────
        ctrl_frame = ttk.Frame(tab2)
        ctrl_frame.grid(row=2, column=0, columnspan=3, pady=6, sticky=tk.W)

        ttk.Label(ctrl_frame, text="Max Hops:").pack(side=tk.LEFT)
        ttk.Spinbox(ctrl_frame, from_=1, to=50, width=5,
                    textvariable=self.targeted_hops_var).pack(side=tk.LEFT, padx=(4, 20))

        self.t_run_btn = ttk.Button(ctrl_frame, text="Run Targeted Search",
                                    command=self.run_targeted_search)
        self.t_run_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.t_stop_btn = ttk.Button(ctrl_frame, text="Stop",
                                     command=self._stop_targeted, state=tk.DISABLED)
        self.t_stop_btn.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(ctrl_frame, text="Clear",
                   command=self._clear_targeted).pack(side=tk.LEFT)

        ttk.Separator(ctrl_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=(18, 10))
        self.graph_btn = ttk.Button(
            ctrl_frame,
            text="📊 Open Graph",
            command=self._open_lineage_graph,
            state=tk.DISABLED,
        )
        self.graph_btn.pack(side=tk.LEFT)

        # ── Row 3: Progress ──────────────────────────────────────────────
        prog_frame = ttk.Frame(tab2)
        prog_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 4))
        prog_frame.columnconfigure(1, weight=1)

        self.t_progress_var = tk.DoubleVar()
        ttk.Progressbar(prog_frame, variable=self.t_progress_var,
                        maximum=100).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        self.t_progress_lbl = ttk.Label(prog_frame, text="", width=14, anchor=tk.E)
        self.t_progress_lbl.grid(row=0, column=2)

        # ── Row 4: Summary label ─────────────────────────────────────────
        self.t_summary_var = tk.StringVar(value="No search run yet.")
        ttk.Label(tab2, textvariable=self.t_summary_var,
                  font=("Consolas", 9), foreground="#1A5276").grid(
            row=4, column=0, columnspan=3, sticky=tk.W, pady=(2, 4))

        # ── Row 5: Export buttons ────────────────────────────────────────
        exp_frame = ttk.Frame(tab2)
        exp_frame.grid(row=5, column=0, columnspan=3, pady=(0, 6), sticky=tk.W)

        ttk.Label(exp_frame, text="Export:").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(exp_frame, text="Column CSV",
                   command=self._export_targeted_column_csv).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(exp_frame, text="Table CSV",
                   command=self._export_targeted_table_csv).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(exp_frame, text="Excel Workbook",
                   command=self._export_targeted_excel).pack(side=tk.LEFT)

        ttk.Separator(exp_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=(12, 8))
        self.ai_doc_btn = ttk.Button(
            exp_frame, text="AI Functional Doc",
            command=self._generate_functional_doc, state=tk.DISABLED,
        )
        self.ai_doc_btn.pack(side=tk.LEFT)

        # ── Row 6: Split pane (tree + log) ───────────────────────────────
        paned = ttk.PanedWindow(tab2, orient=tk.HORIZONTAL)
        paned.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Left: Upstream Lineage Treeview
        tree_lf = ttk.LabelFrame(paned, text="Upstream Lineage Tree", padding=4)
        paned.add(tree_lf, weight=3)
        tree_lf.rowconfigure(0, weight=1)
        tree_lf.columnconfigure(0, weight=1)

        cols = ("Hop", "Type", "Job", "Workflow", "Dataflow")
        self.t_tree = ttk.Treeview(tree_lf, columns=cols, show="tree headings",
                                   selectmode="browse")
        self.t_tree.heading("#0",        text="Table Name")
        self.t_tree.heading("Hop",       text="Hop")
        self.t_tree.heading("Type",      text="Type")
        self.t_tree.heading("Job",       text="Job")
        self.t_tree.heading("Workflow",  text="Workflow")
        self.t_tree.heading("Dataflow",  text="Dataflow")

        self.t_tree.column("#0",        width=220, minwidth=140)
        self.t_tree.column("Hop",       width=42,  minwidth=40,  anchor=tk.CENTER)
        self.t_tree.column("Type",      width=118, minwidth=90)
        self.t_tree.column("Job",       width=155, minwidth=100)
        self.t_tree.column("Workflow",  width=155, minwidth=100)
        self.t_tree.column("Dataflow",  width=170, minwidth=100)

        self.t_tree.tag_configure("target",       background="#2471A3",
                                  foreground="white",
                                  font=("Consolas", 9, "bold"))
        self.t_tree.tag_configure("intermediate", background="#FAD7A0",
                                  foreground="#784212")
        self.t_tree.tag_configure("source",       background="#1E8449",
                                  foreground="white",
                                  font=("Consolas", 9, "bold"))

        tv_sb_v = ttk.Scrollbar(tree_lf, orient=tk.VERTICAL,   command=self.t_tree.yview)
        tv_sb_h = ttk.Scrollbar(tree_lf, orient=tk.HORIZONTAL, command=self.t_tree.xview)
        self.t_tree.configure(yscrollcommand=tv_sb_v.set, xscrollcommand=tv_sb_h.set)
        self.t_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tv_sb_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        tv_sb_h.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Right: Execution log
        log_lf = ttk.LabelFrame(paned, text="Execution Log", padding=4)
        paned.add(log_lf, weight=1)
        log_lf.rowconfigure(0, weight=1)
        log_lf.columnconfigure(0, weight=1)

        self.t_log = scrolledtext.ScrolledText(log_lf, height=20, wrap=tk.WORD,
                                               font=("Consolas", 8))
        self.t_log.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # ── Targeted: helpers ────────────────────────────────────────────────

    def _browse_raw_dump(self):
        """Browse for an optional raw dump XML file (used as column-level fallback)."""
        current = self.raw_dump_path_var.get()
        init_dir = os.path.dirname(current) if os.path.isfile(current) else (
            current if os.path.isdir(current) else os.getcwd()
        )
        path = filedialog.askopenfilename(
            title="Select Raw Dump XML (optional)",
            initialdir=init_dir,
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
        )
        if path:
            self.raw_dump_path_var.set(path)

    def _browse_targeted_output(self):
        current = self.targeted_output_var.get()
        dir_path = filedialog.askdirectory(
            title="Select Output Directory",
            initialdir=current if os.path.isdir(current) else os.path.expanduser("~/Documents"),
        )
        if dir_path:
            self.targeted_output_var.set(dir_path)

    def _browse_targeted_xml(self):
        current = self.targeted_xml_var.get()
        init_dir = os.path.dirname(current) if os.path.isfile(current) else (
            current if os.path.isdir(current) else os.getcwd()
        )
        path = filedialog.askopenfilename(
            title="Select BODS Repository XML",
            initialdir=init_dir,
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
        )
        if path:
            self.targeted_xml_var.set(path)

    def _copy_main_path(self):
        """Copy the Full Extraction input path into the Targeted XML field."""
        p = self.input_path.get().strip()
        if p:
            self.targeted_xml_var.set(p)

    def _t_log_msg(self, msg: str, level: str = "info"):
        """Write a message to the targeted log widget (safe from any thread)."""
        prefix = {"status": "[STATUS]", "info": "[INFO]", "warning": "[WARN]",
                  "error": "[ERROR]", "success": "[OK]"}.get(level, "[INFO]")
        self.t_log.after(0, self._t_log_append, f"{prefix} {msg}\n")

    def _t_log_append(self, text: str):
        self.t_log.insert(tk.END, text)
        self.t_log.see(tk.END)

    # ── Targeted: run ────────────────────────────────────────────────────

    def run_targeted_search(self):
        """Validate inputs and start the targeted lineage worker thread."""
        if self.targeted_is_running:
            return

        xml_path  = self.targeted_xml_var.get().strip()
        raw_input = self.targeted_table_var.get().strip()
        max_hops  = self.targeted_hops_var.get()

        if not xml_path or not os.path.isfile(xml_path):
            messagebox.showerror(
                "Input Folder Not Set",
                f"BODS_FULL_PROD.xml was not found.\n\n"
                f"Browse to the input folder (above the tabs) that contains:\n"
                f"  • {self._MAIN_XML}\n"
                f"  • {self._RAW_DF}  (optional)\n"
                f"  • {self._RAW_ABAP}  (optional)",
                parent=self.root,
            )
            return
        if not raw_input:
            messagebox.showerror("Error", "Enter one or more target table names (comma-separated).", parent=self.root)
            return

        # Parse comma-separated table list
        target_tables = [t.strip() for t in raw_input.split(",") if t.strip()]

        # Reset UI
        self.t_log.delete(1.0, tk.END)
        for item in self.t_tree.get_children():
            self.t_tree.delete(item)
        n = len(target_tables)
        self.t_summary_var.set(f"Running {n} table{'s' if n>1 else ''}…")
        self.t_progress_var.set(0)
        self.t_progress_lbl.config(text="")

        self.targeted_results = []
        self.targeted_is_running = True
        self._targeted_start_time = time.time()
        self.t_run_btn.config(state=tk.DISABLED)
        self.t_stop_btn.config(state=tk.NORMAL)
        if self.graph_btn:
            self.graph_btn.config(state=tk.DISABLED)

        t = threading.Thread(
            target=self._run_targeted_thread,
            args=(xml_path, target_tables, max_hops),
            daemon=True,
        )
        t.start()
        self.root.after(100, self._check_targeted_status)

    def _run_targeted_thread(self, xml_path: str, target_tables: list, max_hops: int):
        """Worker thread — runs TargetedLineageRunner for each table sequentially."""
        try:
            from ds_engine.targeted_lineage_runner import TargetedLineageRunner

            def cb(msg, level="info"):
                self.targeted_queue.put(("log", (msg, level)))
                if level == "status":
                    self.targeted_queue.put(("status", msg))

            # Build RawDumpStore from all raw dump files found in the input folder
            raw_dump = None
            rd_paths = [p for p in self.resolved_raw_dump_paths if os.path.isfile(p)]
            if rd_paths:
                try:
                    from ds_engine.ds_raw_dump_index import RawDumpStore
                    raw_dump = RawDumpStore(rd_paths, status_callback=cb)
                    raw_dump.build_index()
                except Exception as exc:
                    self.targeted_queue.put((
                        "log",
                        (f"[RawDump] Failed to index raw dumps: {exc}", "warning"),
                    ))
                    raw_dump = None

            n = len(target_tables)
            # One runner parses the XML once per table (~40 s each)
            for idx, tbl in enumerate(target_tables):
                if not self.targeted_is_running:
                    break
                pct_start = 5.0 + (idx / n) * 90.0
                self.targeted_queue.put(("progress", pct_start))
                self.targeted_queue.put(("log", (f"[{idx+1}/{n}] Searching: {tbl}", "status")))
                runner = TargetedLineageRunner(xml_path, status_callback=cb)
                result = runner.run(tbl, max_hops=max_hops, raw_dump=raw_dump)
                self.targeted_queue.put(("result", result))
                self.targeted_queue.put(("progress", 5.0 + ((idx + 1) / n) * 90.0))
        except Exception as exc:
            import traceback
            self.targeted_queue.put(("log", (f"FATAL: {exc}", "error")))
            for line in traceback.format_exc().splitlines():
                if line.strip():
                    self.targeted_queue.put(("log", (line, "error")))
        finally:
            self.targeted_queue.put(("done", None))

    def _check_targeted_status(self):
        """Poll the targeted queue and update UI — called via root.after."""
        try:
            while True:
                msg_type, payload = self.targeted_queue.get_nowait()
                if msg_type == "log":
                    msg, level = payload
                    self._t_log_msg(msg, level)
                elif msg_type == "status":
                    self.t_summary_var.set(payload)
                elif msg_type == "progress":
                    pct = float(payload)
                    self.t_progress_var.set(pct)
                    self.t_progress_lbl.config(text=f"{pct:.1f}%")
                elif msg_type == "result":
                    self.targeted_results.append(payload)
                elif msg_type == "done":
                    self._targeted_finished()
                    return
        except queue.Empty:
            pass

        if self.targeted_is_running:
            self.root.after(100, self._check_targeted_status)

    def _targeted_finished(self):
        """Called when the targeted worker thread signals done."""
        self.targeted_is_running = False
        self.t_run_btn.config(state=tk.NORMAL)
        self.t_stop_btn.config(state=tk.DISABLED)
        self.t_progress_var.set(100)
        self.t_progress_lbl.config(text="100.0%")

        elapsed = time.time() - self._targeted_start_time
        m, s = divmod(int(elapsed), 60)
        elapsed_str = f"{m}m {s}s" if m else f"{s}s"

        if self.targeted_results:
            n = len(self.targeted_results)
            total_rows  = sum(r.get("stats", {}).get("total_rows", 0)  for r in self.targeted_results)
            total_terms = sum(r.get("stats", {}).get("terminal_srcs", 0) for r in self.targeted_results)
            tables_str  = ", ".join(r["target_table"] for r in self.targeted_results)
            self.t_summary_var.set(
                f"{n} table{'s' if n>1 else ''}: {tables_str}  |  "
                f"{total_terms} terminal source(s)  |  {total_rows} column row(s)  |  {elapsed_str}"
            )
            # Clear once, then append each result (fixes per-call wipe losing earlier tables)
            for item in self.t_tree.get_children():
                self.t_tree.delete(item)
            for result in self.targeted_results:
                self._populate_targeted_tree(result)
            self._populate_col_tree()
            if self.graph_btn:
                self.graph_btn.config(state=tk.NORMAL)
            if hasattr(self, "ai_doc_btn"):
                self.ai_doc_btn.config(state=tk.NORMAL)
            # Prepend to recent-searches combobox (deduplicated, capped at 15)
            search_str = self.targeted_table_var.get().strip()
            if search_str and (not self.targeted_recent_tables or self.targeted_recent_tables[0] != search_str):
                self.targeted_recent_tables = (
                    [search_str] + [x for x in self.targeted_recent_tables if x != search_str]
                )[:15]
                self.t_table_entry['values'] = self.targeted_recent_tables
        else:
            self.t_summary_var.set(
                f"Search finished — no result (check the log for errors).  |  {elapsed_str}"
            )
        self._save_session_state()

    def _stop_targeted(self):
        """Stop the targeted search (marks flag; thread will finish naturally)."""
        if self.targeted_is_running:
            self.targeted_is_running = False
            self.targeted_queue.put(("log", ("Search stopped by user.", "warning")))
            self._targeted_finished()

    def _clear_targeted(self):
        self.t_log.delete(1.0, tk.END)
        for item in self.t_tree.get_children():
            self.t_tree.delete(item)
        self.t_summary_var.set("No search run yet.")
        self.t_progress_var.set(0)
        self.t_progress_lbl.config(text="")
        self.targeted_results = []
        if hasattr(self, "ai_doc_btn"):
            self.ai_doc_btn.config(state=tk.DISABLED)

    # ── Targeted: tree population ─────────────────────────────────────────

    def _populate_targeted_tree(self, result: dict):
        """Append one target's lineage subtree to the targeted tree (no clear)."""
        if not result:
            return

        target_upper  = result["target_table"].upper()
        upstream_tree = result.get("upstream_tree", {})
        hop_map       = result.get("hop_map", {})
        terminals     = {t.upper() for t in result.get("terminal_sources", set())}
        # Prefer table_context (built from FlatExportIndex) over column-lineage rows
        table_context = result.get("table_context", {})

        def add_node(parent_id, table: str, path: frozenset):
            tup = table.upper()
            if tup in path:           # cycle guard
                return
            hop = hop_map.get(tup, "?")
            if hop == 0:
                node_type, tag = "TARGET",       "target"
            elif tup in terminals:
                node_type, tag = "FINAL SOURCE", "source"
            else:
                node_type, tag = "INTERMEDIATE", "intermediate"

            ctx = table_context.get(tup, {})
            values = (str(hop), node_type,
                      ctx.get("Job_Name", ""),
                      ctx.get("Workflow_Name", ""),
                      ctx.get("Dataflow_Name", ""))

            item_id = self.t_tree.insert(
                parent_id, "end",
                text=table,
                values=values,
                open=(isinstance(hop, int) and hop < 2),
                tags=(tag,),
            )
            for child in upstream_tree.get(tup, []):
                add_node(item_id, child, path | {tup})

        add_node("", target_upper, frozenset())

    # ── Targeted: export ─────────────────────────────────────────────────

    def _export_targeted_column_csv(self):
        if not self.targeted_results:
            messagebox.showwarning("No Data", "Run a targeted search first.", parent=self.root)
            return
        all_rows = [r for res in self.targeted_results for r in res.get("rows", [])]
        if not all_rows:
            messagebox.showwarning("No Data", "No column-level rows to export.", parent=self.root)
            return
        safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", self.targeted_results[0]["target_table"])
        out_dir = self.targeted_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, f"targeted_column_{safe_name}.csv")
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Column Lineage CSV",
                defaultextension=".csv",
                initialfile=f"targeted_column_{safe_name}.csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            pd.DataFrame(all_rows).to_csv(fp, index=False)
            messagebox.showinfo("Exported", f"Column lineage saved:\n{fp}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)

    def _export_targeted_table_csv(self):
        if not self.targeted_results:
            messagebox.showwarning("No Data", "Run a targeted search first.", parent=self.root)
            return
        safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", self.targeted_results[0]["target_table"])
        out_dir = self.targeted_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, f"targeted_table_{safe_name}.csv")
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Table Lineage CSV",
                defaultextension=".csv",
                initialfile=f"targeted_table_{safe_name}.csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            records = []
            for result in self.targeted_results:
                hop_map       = result.get("hop_map", {})
                upstream_tree = result.get("upstream_tree", {})
                rows          = result.get("rows", [])
                tbl_ctx: dict = {}
                for r in rows:
                    tgt = (r.get("Target_Object") or "").upper()
                    if tgt and tgt not in tbl_ctx:
                        tbl_ctx[tgt] = {k: r.get(k, "") for k in ("Job_Name","Workflow_Name","Dataflow_Name")}
                for tgt, srcs in upstream_tree.items():
                    ctx = tbl_ctx.get(tgt.upper(), {})
                    for src in srcs:
                        records.append({
                            "Target_Table":  tgt,
                            "Source_Table":  src,
                            "Target_Hop":    hop_map.get(tgt.upper(), ""),
                            "Source_Hop":    hop_map.get(src.upper(), ""),
                            "Job_Name":      ctx.get("Job_Name", ""),
                            "Workflow_Name": ctx.get("Workflow_Name", ""),
                            "Dataflow_Name": ctx.get("Dataflow_Name", ""),
                        })
            if not records:
                messagebox.showwarning("No Data", "No table lineage data to export.", parent=self.root)
                return
            pd.DataFrame(records).to_csv(fp, index=False)
            messagebox.showinfo("Exported", f"Table lineage saved:\n{fp}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)

    def _export_targeted_excel(self):
        if not self.targeted_results:
            messagebox.showwarning("No Data", "Run a targeted search first.", parent=self.root)
            return
        n         = len(self.targeted_results)
        safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", self.targeted_results[0]["target_table"])
        default_name = f"targeted_lineage_{safe_name}.xlsx" if n == 1 else f"targeted_lineage_multi_{safe_name}.xlsx"
        out_dir = self.targeted_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, default_name)
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Excel Workbook",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            from datetime import datetime as _dt

            with pd.ExcelWriter(fp, engine="openpyxl") as writer:
                # Summary sheet — one row per target
                summary_rows = []
                for result in self.targeted_results:
                    stats = result.get("stats", {})
                    summary_rows.append({
                        "Target_Table":     result["target_table"],
                        "Total_Column_Rows": stats.get("total_rows", ""),
                        "Hops":             stats.get("hops", ""),
                        "Dataflows":        stats.get("dataflows", ""),
                        "Terminal_Sources": stats.get("terminal_srcs", ""),
                        "Jobs":             stats.get("jobs", ""),
                        "Elapsed_s":        stats.get("elapsed_s", ""),
                        "Generated":        _dt.now().strftime("%Y-%m-%d %H:%M"),
                    })
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

                for result in self.targeted_results:
                    hop_map       = result.get("hop_map", {})
                    upstream_tree = result.get("upstream_tree", {})
                    terminals     = {t.upper() for t in result.get("terminal_sources", set())}
                    table_context = result.get("table_context", {})
                    rows          = result.get("rows", [])

                    table_records = []
                    for tgt, srcs in upstream_tree.items():
                        ctx      = table_context.get(tgt.upper(), table_context.get(tgt, {}))
                        tgt_hop  = hop_map.get(tgt.upper(), hop_map.get(tgt, ""))
                        tgt_type = "TARGET" if tgt_hop == 0 else "INTERMEDIATE"
                        for src in srcs:
                            src_hop  = hop_map.get(src.upper(), hop_map.get(src, ""))
                            src_type = "FINAL SOURCE" if src.upper() in terminals else "INTERMEDIATE"
                            table_records.append({
                                "Lineage_Path":  f"{src}  -->  {tgt}  (via {ctx.get('Dataflow_Name','?').split(';')[0].strip()} | {ctx.get('Job_Name','?').split(';')[0].strip()})",
                                "Target_Table":  tgt, "Target_Type": tgt_type, "Target_Hop": tgt_hop,
                                "Source_Table":  src, "Source_Type": src_type, "Source_Hop": src_hop,
                                "Job_Name":      ctx.get("Job_Name", ""),
                                "Workflow_Name": ctx.get("Workflow_Name", ""),
                                "Dataflow_Name": ctx.get("Dataflow_Name", ""),
                            })

                    safe_sheet = re.sub(r"[^A-Za-z0-9_]", "_", result["target_table"])[:28]
                    col_df = pd.DataFrame(rows) if rows else pd.DataFrame()
                    if not col_df.empty:
                        col_df.to_excel(writer, sheet_name=f"{safe_sheet}_Col", index=False)
                    if table_records:
                        tdf = pd.DataFrame(table_records)
                        _smap = {"TARGET": 0, "INTERMEDIATE": 1, "FINAL SOURCE": 2}
                        tdf["_s"] = tdf["Source_Type"].map(_smap).fillna(1)
                        tdf.sort_values(["Target_Hop", "Source_Hop", "_s"], inplace=True)
                        tdf.drop(columns=["_s"], inplace=True)
                        tdf.reset_index(drop=True, inplace=True)
                        tdf.to_excel(writer, sheet_name=f"{safe_sheet}_Tbl", index=False)
                        self._format_table_lineage_sheet(writer.sheets[f"{safe_sheet}_Tbl"], tdf)

            messagebox.showinfo("Exported", f"Excel workbook saved:\n{fp}", parent=self.root)
        except ImportError:
            messagebox.showerror("Export Error", "openpyxl not installed. Run: pip install openpyxl", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)

    # ── AI Functional Document ────────────────────────────────────────────────

    def _generate_functional_doc(self):
        """Generate a low-level AI functional specification document (.docx)."""
        if not self.targeted_results:
            messagebox.showwarning("No Data", "Run a targeted search first.", parent=self.root)
            return

        api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
        if not api_key:
            messagebox.showerror(
                "API Key Missing",
                "Set the ANTHROPIC_API_KEY environment variable and restart the app.\n\n"
                "Example (PowerShell):\n  $env:ANTHROPIC_API_KEY = 'sk-ant-...'",
                parent=self.root,
            )
            return

        safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", self.targeted_results[0]["target_table"])
        ts        = time.strftime("%Y%m%d_%H%M%S")
        out_dir   = self.targeted_output_var.get().strip()

        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, f"FuncSpec_{safe_name}_{ts}.docx")
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Functional Specification",
                defaultextension=".docx",
                initialfile=f"FuncSpec_{safe_name}_{ts}.docx",
                filetypes=[("Word Document", "*.docx"), ("All files", "*.*")],
                parent=self.root,
            )
        if not fp:
            return

        self.ai_doc_btn.config(state=tk.DISABLED, text="Generating…")

        results_snapshot = list(self.targeted_results)

        def _log(msg: str):
            self.root.after(0, lambda m=msg: (
                self.t_log.insert(tk.END, m + "\n"),
                self.t_log.see(tk.END),
            ))

        def _worker():
            try:
                from ds_engine.ds_ai_doc_generator import generate_functional_doc
                generate_functional_doc(
                    results=results_snapshot,
                    output_path=fp,
                    api_key=api_key,
                    status_callback=_log,
                )
                self.root.after(0, lambda: messagebox.showinfo(
                    "Done",
                    f"Functional specification saved:\n{fp}",
                    parent=self.root,
                ))
            except ImportError as exc:
                self.root.after(0, lambda e=exc: messagebox.showerror(
                    "Missing Library",
                    f"Install required packages:\n  pip install anthropic python-docx\n\n{e}",
                    parent=self.root,
                ))
            except Exception as exc:
                self.root.after(0, lambda e=exc: messagebox.showerror(
                    "Generation Error", str(e), parent=self.root,
                ))
            finally:
                self.root.after(0, lambda: self.ai_doc_btn.config(
                    state=tk.NORMAL, text="AI Functional Doc"
                ))

        threading.Thread(target=_worker, daemon=True).start()

    def _format_table_lineage_sheet(self, ws, df):
        """
        Apply colour coding, bold header, auto-fit columns and freeze pane
        to the Table_Lineage worksheet so non-technical users can read it easily.

        Row colours:
          TARGET rows     → light blue   (#D6EAF8)
          FINAL SOURCE    → light green  (#D5F5E3)
          INTERMEDIATE    → light yellow (#FEF9E7)
        """
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return   # openpyxl not available — skip formatting silently

        # Colour fills
        HDR_FILL  = PatternFill("solid", fgColor="1B2631")  # near-black header
        TGT_FILL  = PatternFill("solid", fgColor="AED6F1")  # medium blue
        INT_FILL  = PatternFill("solid", fgColor="FEF9E7")  # pale cream
        SRC_FILL  = PatternFill("solid", fgColor="A9DFBF")  # soft green
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
        THIN      = Side(style="thin", color="BDC3C7")
        BORDER    = Border(bottom=THIN)

        col_count = len(df.columns)

        # Header row
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        src_type_col = None
        tgt_type_col = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == "Source_Type": src_type_col = idx
            if col_name == "Target_Type": tgt_type_col = idx

        for row_idx in range(2, len(df) + 2):
            src_type = ws.cell(row=row_idx, column=src_type_col).value if src_type_col else ""
            tgt_type = ws.cell(row=row_idx, column=tgt_type_col).value if tgt_type_col else ""
            if src_type == "FINAL SOURCE":
                fill = SRC_FILL
            elif tgt_type == "TARGET":
                fill = TGT_FILL
            else:
                fill = INT_FILL
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Auto-fit column widths (cap at 60 chars)
        for col_idx, col_name in enumerate(df.columns, 1):
            col_vals = [str(col_name)] + [
                str(v) if v is not None else "" for v in df.iloc[:, col_idx - 1]
            ]
            best_width = min(max(len(v) for v in col_vals) + 3, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = best_width

        # Freeze header row and set row height
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def _open_lineage_graph(self):
        """Build the multi-tab HTML graph and open it in the system browser."""
        if not self.targeted_results:
            return
        import webbrowser
        try:
            path = self._build_lineage_html(self.targeted_results)
            webbrowser.open("file:///" + path.replace(os.sep, "/"))
        except Exception as exc:
            messagebox.showerror("Graph Error", str(exc), parent=self.root)

    @staticmethod
    def _result_to_records(result: dict) -> list:
        """Convert a TargetedLineageRunner result dict to a flat records list for the HTML template."""
        hop_map       = result.get("hop_map", {})
        upstream_tree = result.get("upstream_tree", {})
        terminals     = {t.upper() for t in result.get("terminal_sources", set())}
        table_context = result.get("table_context", {})
        records = []
        for tgt_tbl, srcs in upstream_tree.items():
            tgt_up  = tgt_tbl.upper()
            ctx     = table_context.get(tgt_up, table_context.get(tgt_tbl, {}))
            tgt_hop = hop_map.get(tgt_up, hop_map.get(tgt_tbl, 0))
            if isinstance(tgt_hop, float):
                tgt_hop = int(tgt_hop)
            for src_tbl in srcs:
                src_up  = src_tbl.upper()
                src_hop = hop_map.get(src_up, hop_map.get(src_tbl, 0))
                if isinstance(src_hop, float):
                    src_hop = int(src_hop)
                records.append({
                    "src":     src_tbl,
                    "src_hop": src_hop,
                    "is_src":  src_up in terminals,
                    "tgt":     tgt_tbl,
                    "tgt_hop": tgt_hop,
                    "job":     ctx.get("Job_Name", ""),
                    "wf":      ctx.get("Workflow_Name", ""),
                    "df":      ctx.get("Dataflow_Name", ""),
                })

        # Inject SQL transform nodes: one synthetic SQL:<instance> → target edge per
        # unique (transform_instance, target) pair found in SQL_TRANSFORM rows.
        # Use Transform_Name (the BODS instance name, e.g. "SQL_IRI") if available,
        # falling back to Dataflow_Name so older rows without it still work.
        # Skip any edge already present from the upstream_tree (Phase E1 BFS route).
        _upstream_sql_edges = {
            (r["src"], r["tgt"]) for r in records if r["src"].startswith("SQL:")
        }
        # Phase E1 compound keys use "SQL:df_name|tf_name"; SQL_TRANSFORM rows use the
        # simple form "SQL:tf_name".  Build a simple-form set so the dedup below can
        # match "SQL:SQL" against the compound key "SQL:OUT_IRI_RESP|SQL".
        _upstream_sql_simple = {
            (f"SQL:{r['src'].split('|')[-1]}", r["tgt"])
            for r in records
            if r["src"].startswith("SQL:") and "|" in r["src"]
        }
        sql_seen: set = set()
        for row in result.get("rows", []):
            if row.get("Record_Type") != "SQL_TRANSFORM":
                continue
            df_name        = (row.get("Dataflow_Name", "") or "SQL").strip()
            transform_name = (row.get("Transform_Name") or df_name or "SQL").strip()
            tgt            = (row.get("Target_Object", "") or "").strip()
            if not tgt or not transform_name:
                continue
            key = (transform_name, tgt)
            if key in sql_seen:
                continue
            sql_seen.add(key)
            sql_node = f"SQL:{transform_name}"
            # Skip if this SQL→target edge was already produced by the upstream_tree
            # (match against both compound "SQL:df|tf" and simple "SQL:tf" forms).
            if (sql_node, tgt) in _upstream_sql_edges or (sql_node, tgt) in _upstream_sql_simple:
                continue
            tgt_up   = tgt.upper()
            # Only emit this synthetic edge if the target is a known physical table.
            # If tgt is an unresolved transform output view it won't be in hop_map,
            # and the default of 0 would falsely place it at the target position.
            if tgt_up not in hop_map:
                continue
            tgt_hop = hop_map[tgt_up]
            if isinstance(tgt_hop, float):
                tgt_hop = int(tgt_hop)
            records.append({
                "src":     sql_node,
                "src_hop": tgt_hop + 1,
                "is_src":  False,
                "tgt":     tgt,
                "tgt_hop": tgt_hop,
                "job":     row.get("Job_Name", ""),
                "wf":      row.get("Workflow_Name", ""),
                "df":      df_name,
            })

        return records

    @staticmethod
    def _result_to_col_data(result: dict) -> dict:
        """
        Build per-target column detail from result['rows'].
        Returns { "TGT_TABLE": {"dataflows": [{df, job, wf, mappings, joins, filters}]} }
        Keys use canonical graph node names (resolved via hop_map).
        """
        rows = result.get("rows", [])

        # Map short/full table names → canonical graph node name (from hop_map)
        hop_map = result.get("hop_map", {})
        _short: dict = {}
        for full in hop_map:
            _short[full.upper()] = full
            short = full.split(".")[-1].upper()
            if short not in _short:
                _short[short] = full

        def norm(name: str) -> str:
            n = (name or "").strip().upper()
            hit = _short.get(n)
            if hit is not None:
                return hit
            # Schema-qualified "SCHEMA.TABLE" → try bare "TABLE" segment so that
            # raw-dump rows with a fully-qualified target resolve to the hop_map
            # key even when the user searched using the bare table name.
            bare = n.split(".")[-1]
            if bare != n:
                hit = _short.get(bare)
                if hit is not None:
                    return hit
            return (name or "").strip()

        # by_tgt: { canonical_tgt: { df_name: {mappings, joins, filters, job, wf} } }
        by_tgt: dict = {}

        _COL_RECORD_TYPES = {
            "COLUMN_MAPPING", "UNION_COLUMN_MAPPING", "SCRIPT_COLUMN_MAPPING",
            "TABLE_COMPARISON", "SQL_TRANSFORM", "REVERSE_PIVOT", "PIVOT",
            "HISTORY_PRESERVING", "KEY_GENERATION", "CASE_OPERATION",
        }

        # Pass 1 — column mapping rows → group by target table then by dataflow
        for row in rows:
            rt = row.get("Record_Type", "")
            if rt not in _COL_RECORD_TYPES:
                continue
            tgt = norm(row.get("Target_Object", ""))
            if not tgt or tgt.upper() not in _short:
                continue  # Target_Object is an unresolved transform output view — skip
            df = row.get("Dataflow_Name", "") or "(no dataflow)"
            if tgt not in by_tgt:
                by_tgt[tgt] = {}
            if df not in by_tgt[tgt]:
                by_tgt[tgt][df] = {
                    "mappings": [], "joins": [], "filters": [],
                    "job": row.get("Job_Name", "") or "",
                    "wf":  row.get("Workflow_Name", "") or "",
                }
            formula = (row.get("Formula") or row.get("Expression_Text") or "").strip()
            is_direct = (row.get("Transformation_Type", "") == "DIRECT_SOURCE"
                         or row.get("Transformation_Category", "") == "PASS_THROUGH")
            by_tgt[tgt][df]["mappings"].append({
                "sc":      row.get("Source_Column", ""),
                "tc":      row.get("Target_Column", ""),
                "f":       formula,
                "d":       is_direct,
                "df":      df,
                # Use Transformation_Type first so tier-1/2 badges display correctly
                "ft":      (row.get("Transformation_Type")
                            or row.get("Formula_Type")
                            or row.get("Formula_Category") or ""),
                "fu":      row.get("Functions_Used", ""),
                "src_obj": row.get("Source_Object", ""),
            })

        # Pass 2 — associate joins and filters to the matching dataflow entries
        for row in rows:
            rt = row.get("Record_Type", "")
            df = row.get("Dataflow_Name", "") or "(no dataflow)"
            if rt in ("JOIN", "JOIN_KEY_MAPPING"):
                entry = {
                    "jt":   row.get("Join_Type", "JOIN"),
                    "lo":   row.get("Join_Left_Object", ""),
                    "ro":   row.get("Join_Right_Object", ""),
                    "lc":   row.get("Join_Left_Column", ""),
                    "rc":   row.get("Join_Right_Column", ""),
                    "cond": row.get("Join_Condition", ""),
                    "df":   df,
                }
                for tgt_data in by_tgt.values():
                    if df in tgt_data and entry not in tgt_data[df]["joins"]:
                        tgt_data[df]["joins"].append(entry)
            elif rt == "FILTER":
                cond = row.get("Where_Condition", "")
                if not cond:
                    continue
                tgt = norm(row.get("Target_Object", ""))
                entry = {"cond": cond, "df": df}
                if tgt in by_tgt and df in by_tgt[tgt]:
                    if entry not in by_tgt[tgt][df]["filters"]:
                        by_tgt[tgt][df]["filters"].append(entry)

        # Inject synthetic context-only entries for tables that produced no column rows.
        # Ensures the correct job / dataflow shows in the panel picker even when the
        # raw dump XML was not provided and DIQuery bodies were not extracted.
        for tbl_up, ctx in result.get("table_context", {}).items():
            canonical = _short.get(tbl_up, tbl_up)
            if canonical in by_tgt:
                continue  # actual column rows already exist — don't add a synthetic stub
            # table_context joins multiple values with "; " — split each df separately
            df_names = [x.strip() for x in (ctx.get("Dataflow_Name", "") or "").split(";")
                        if x.strip()]
            if not df_names:
                continue
            by_tgt[canonical] = {
                df_nm: {
                    "mappings": [],
                    "joins":    [],
                    "filters":  [],
                    "job":      ctx.get("Job_Name", "") or "",
                    "wf":       ctx.get("Workflow_Name", "") or "",
                }
                for df_nm in df_names
            }

        # Build final output: { tgt: {"dataflows": [...]} }
        col_data: dict = {}
        for tgt, df_map in by_tgt.items():
            col_data[tgt] = {
                "dataflows": [
                    {
                        "df":       df_name,
                        "job":      df_data["job"],
                        "wf":       df_data["wf"],
                        "mappings": df_data["mappings"],
                        "joins":    df_data["joins"],
                        "filters":  df_data["filters"],
                    }
                    for df_name, df_data in df_map.items()
                ]
            }
        return col_data

    def _build_lineage_html(self, results) -> str:
        """
        Generate a self-contained multi-tab interactive HTML lineage graph.
        Accepts a list of result dicts (one tab per entry) or a single result dict.
        Returns the path of the written HTML file.
        """
        import json

        if isinstance(results, dict):
            results = [results]

        datasets = []
        for result in results:
            records = self._result_to_records(result)
            if records:
                col_data = self._result_to_col_data(result)
                datasets.append({
                    "tgt":         result["target_table"],
                    "recs":        records,
                    "cols":        col_data,
                    "sql_members": result.get("sql_members", {}),
                    # Include both compound keys ("SQL:df|tf") and simple aliases
                    # ("SQL:tf") so showDetail finds the query regardless of which
                    # node form ends up in the graph after dedup.
                    "sql_queries": {
                        # Fallback: SQL_TRANSFORM rows always carry Formula=sql_text
                        # (produced by ds_transform_extractor second pass).  Use as
                        # base so the panel shows SQL even when Phase E1 streaming
                        # missed the sql_text element in the XML.
                        **{
                            f"SQL:{(row.get('Transform_Name') or row.get('Dataflow_Name') or 'SQL').strip()}": (row.get("Formula") or "").strip()
                            for row in result.get("rows", [])
                            if row.get("Record_Type") == "SQL_TRANSFORM"
                            and (row.get("Formula") or "").strip()
                        },
                        # Phase E1 compound keys override (more specific)
                        **result.get("sql_queries", {}),
                        # Simple-key aliases for compound keys so JS lookup by
                        # tfName still works ("SQL:SQL" → same text as "SQL:df|SQL")
                        **{
                            f"SQL:{ck[4:].split('|')[-1]}": qt
                            for ck, qt in result.get("sql_queries", {}).items()
                            if "|" in ck[4:]
                        },
                    },
                })

        if not datasets:
            raise ValueError(
                "No lineage edges found for any of the requested tables. "
                "Run the targeted search first."
            )

        html = _LINEAGE_HTML_TEMPLATE.replace("%%DATASETS%%", json.dumps(datasets))

        # Output path alongside the XML file.
        # Timestamp suffix ensures each run gets a unique filename so browsers
        # always open a fresh tab instead of reusing a cached local-file URL.
        import time as _time
        first_safe = re.sub(r"[^A-Za-z0-9_.-]", "_", results[0]["target_table"])
        suffix     = f"_and_{len(results)-1}_more" if len(results) > 1 else ""
        fname      = f"lineage_{first_safe[:40]}{suffix}_{int(_time.time())}.html"
        out_dir = self.targeted_output_var.get().strip()
        if not out_dir or not os.path.isdir(out_dir):
            xml_path = self.targeted_xml_var.get() or self.xml_var.get() or ""
            if xml_path and os.path.isfile(xml_path):
                out_dir = os.path.dirname(os.path.abspath(xml_path))
            else:
                out_dir = os.path.expanduser(os.path.join("~", "Documents"))
                os.makedirs(out_dir, exist_ok=True)

        out_path = os.path.join(out_dir, fname)
        with open(out_path, "w", encoding="utf-8") as fh:
            fh.write(html)
        return out_path


    # ======================================================================
    # Job Level Tab (Tab 3)
    # ======================================================================

    def _create_job_tab(self):
        """Build the Job Level Analysis tab (Tab 3)."""
        tab3 = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab3, text="  Job Level  ")

        tab3.columnconfigure(1, weight=1)
        tab3.rowconfigure(6, weight=1)

        # ── Row 0: Output directory ──────────────────────────────────────
        ttk.Label(tab3, text="Output Dir:").grid(row=0, column=0, sticky=tk.W, pady=4)
        self.j_output_entry = ttk.Entry(tab3, textvariable=self.job_output_var, width=55)
        self.j_output_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=4)
        ttk.Button(tab3, text="Browse…",
                   command=self._browse_job_output).grid(row=0, column=2, padx=(5, 0), pady=4, sticky=tk.W)

        # ── Row 1: Job name(s) ───────────────────────────────────────────
        ttk.Label(tab3, text="Job Name(s):").grid(row=1, column=0, sticky=tk.W, pady=4)
        self.j_name_entry = ttk.Combobox(tab3, textvariable=self.job_name_var, width=55)
        self.j_name_entry['values'] = self.job_recent_tables
        self.j_name_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 5), pady=4)
        ttk.Label(tab3, text="comma-separated for multiple jobs", foreground="grey").grid(
            row=1, column=2, sticky=tk.W, padx=(5, 0))

        # ── Row 2: Options + controls ────────────────────────────────────
        ctrl_frame = ttk.Frame(tab3)
        ctrl_frame.grid(row=2, column=0, columnspan=3, pady=6, sticky=tk.W)

        ttk.Label(ctrl_frame, text="Max Hops:").pack(side=tk.LEFT)
        ttk.Spinbox(ctrl_frame, from_=1, to=50, width=5,
                    textvariable=self.job_hops_var).pack(side=tk.LEFT, padx=(4, 20))

        self.j_run_btn = ttk.Button(ctrl_frame, text="Run Job Analysis",
                                    command=self.run_job_search)
        self.j_run_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.j_stop_btn = ttk.Button(ctrl_frame, text="Stop",
                                     command=self._stop_job, state=tk.DISABLED)
        self.j_stop_btn.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(ctrl_frame, text="Clear",
                   command=self._clear_job).pack(side=tk.LEFT)

        ttk.Separator(ctrl_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=(18, 10))
        self.job_graph_btn = ttk.Button(
            ctrl_frame,
            text="📊 Open Graph",
            command=self._open_job_graph,
            state=tk.DISABLED,
        )
        self.job_graph_btn.pack(side=tk.LEFT)

        # ── Row 3: Progress ──────────────────────────────────────────────
        prog_frame = ttk.Frame(tab3)
        prog_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 4))
        prog_frame.columnconfigure(1, weight=1)

        self.j_progress_var = tk.DoubleVar()
        ttk.Progressbar(prog_frame, variable=self.j_progress_var,
                        maximum=100).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 5))
        self.j_progress_lbl = ttk.Label(prog_frame, text="", width=14, anchor=tk.E)
        self.j_progress_lbl.grid(row=0, column=2)

        # ── Row 4: Summary label ─────────────────────────────────────────
        self.j_summary_var = tk.StringVar(value="No job analysis run yet.")
        ttk.Label(tab3, textvariable=self.j_summary_var,
                  font=("Consolas", 9), foreground="#1A5276").grid(
            row=4, column=0, columnspan=3, sticky=tk.W, pady=(2, 4))

        # ── Row 5: Export buttons ────────────────────────────────────────
        exp_frame = ttk.Frame(tab3)
        exp_frame.grid(row=5, column=0, columnspan=3, pady=(0, 6), sticky=tk.W)

        ttk.Label(exp_frame, text="Export:").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(exp_frame, text="Column CSV",
                   command=self._export_job_column_csv).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(exp_frame, text="Table CSV",
                   command=self._export_job_table_csv).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(exp_frame, text="Excel Workbook",
                   command=self._export_job_excel).pack(side=tk.LEFT)

        # ── Row 6: Split pane (tree + log) ───────────────────────────────
        paned = ttk.PanedWindow(tab3, orient=tk.HORIZONTAL)
        paned.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))

        tree_lf = ttk.LabelFrame(paned, text="Job Lineage Tree", padding=4)
        paned.add(tree_lf, weight=3)
        tree_lf.rowconfigure(0, weight=1)
        tree_lf.columnconfigure(0, weight=1)

        cols = ("Hop", "Type", "Job", "Workflow", "Dataflow")
        self.j_tree = ttk.Treeview(tree_lf, columns=cols, show="tree headings",
                                   selectmode="browse")
        self.j_tree.heading("#0",        text="Table Name")
        self.j_tree.heading("Hop",       text="Hop")
        self.j_tree.heading("Type",      text="Type")
        self.j_tree.heading("Job",       text="Job")
        self.j_tree.heading("Workflow",  text="Workflow")
        self.j_tree.heading("Dataflow",  text="Dataflow")

        self.j_tree.column("#0",        width=220, minwidth=140)
        self.j_tree.column("Hop",       width=42,  minwidth=40,  anchor=tk.CENTER)
        self.j_tree.column("Type",      width=118, minwidth=90)
        self.j_tree.column("Job",       width=155, minwidth=100)
        self.j_tree.column("Workflow",  width=155, minwidth=100)
        self.j_tree.column("Dataflow",  width=170, minwidth=100)

        self.j_tree.tag_configure("target",       background="#2471A3",
                                  foreground="white",
                                  font=("Consolas", 9, "bold"))
        self.j_tree.tag_configure("intermediate", background="#FAD7A0",
                                  foreground="#784212")
        self.j_tree.tag_configure("source",       background="#1E8449",
                                  foreground="white",
                                  font=("Consolas", 9, "bold"))
        self.j_tree.tag_configure("job_root",     background="#4A235A",
                                  foreground="white",
                                  font=("Consolas", 9, "bold"))

        jv_sb_v = ttk.Scrollbar(tree_lf, orient=tk.VERTICAL,   command=self.j_tree.yview)
        jv_sb_h = ttk.Scrollbar(tree_lf, orient=tk.HORIZONTAL, command=self.j_tree.xview)
        self.j_tree.configure(yscrollcommand=jv_sb_v.set, xscrollcommand=jv_sb_h.set)
        self.j_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        jv_sb_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        jv_sb_h.grid(row=1, column=0, sticky=(tk.W, tk.E))

        log_lf = ttk.LabelFrame(paned, text="Execution Log", padding=4)
        paned.add(log_lf, weight=1)
        log_lf.rowconfigure(0, weight=1)
        log_lf.columnconfigure(0, weight=1)

        self.j_log = scrolledtext.ScrolledText(log_lf, height=20, wrap=tk.WORD,
                                               font=("Consolas", 8))
        self.j_log.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # ── Column Lineage tab ────────────────────────────────────────────────

    def _create_column_tab(self):
        """Build the Column Lineage tab (Tab 4) — field-level mapping viewer."""
        tab4 = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab4, text="  Column Lineage  ")

        tab4.columnconfigure(0, weight=1)
        tab4.rowconfigure(2, weight=1)

        # Row 0: summary + export button
        top_frame = ttk.Frame(tab4)
        top_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 4))
        top_frame.columnconfigure(1, weight=1)

        self.col_summary_var = tk.StringVar(
            value="Run a Targeted Search (Tab 2) to see column-level field mappings here.")
        ttk.Label(top_frame, textvariable=self.col_summary_var,
                  font=("Consolas", 9), foreground="#1A5276").grid(
            row=0, column=0, sticky=tk.W)
        self.col_export_btn = ttk.Button(
            top_frame, text="Export CSV",
            command=self._export_col_csv, state=tk.DISABLED)
        self.col_export_btn.grid(row=0, column=2, sticky=tk.E, padx=(10, 0))

        # Row 1: filter bar
        flt_frame = ttk.Frame(tab4)
        flt_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 4))
        flt_frame.columnconfigure(1, weight=1)
        ttk.Label(flt_frame, text="Filter:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.col_filter_var = tk.StringVar()
        self.col_filter_var.trace_add("write", lambda *_: self._apply_col_filter())
        ttk.Entry(flt_frame, textvariable=self.col_filter_var, width=40).grid(
            row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        ttk.Label(flt_frame, text="table or column name", foreground="grey").grid(
            row=0, column=2, sticky=tk.W)

        # Row 2: treeview
        tree_fr = ttk.Frame(tab4)
        tree_fr.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_fr.rowconfigure(0, weight=1)
        tree_fr.columnconfigure(0, weight=1)

        cols = ("src_tbl", "expression", "tgt_col", "dataflow", "job")
        self.col_tree = ttk.Treeview(tree_fr, columns=cols, show="tree headings",
                                     selectmode="browse")
        self.col_tree.heading("#0",         text="Source Column  →  Target Column")
        self.col_tree.heading("src_tbl",    text="Source Table")
        self.col_tree.heading("expression", text="Expression")
        self.col_tree.heading("tgt_col",    text="Target Column")
        self.col_tree.heading("dataflow",   text="Dataflow")
        self.col_tree.heading("job",        text="Job")

        self.col_tree.column("#0",         width=270, minwidth=160)
        self.col_tree.column("src_tbl",    width=200, minwidth=120)
        self.col_tree.column("expression", width=200, minwidth=100)
        self.col_tree.column("tgt_col",    width=140, minwidth=80)
        self.col_tree.column("dataflow",   width=160, minwidth=100)
        self.col_tree.column("job",        width=140, minwidth=100)

        self.col_tree.tag_configure(
            "target_hdr", background="#2471A3", foreground="white",
            font=("Consolas", 9, "bold"))
        self.col_tree.tag_configure("mapping", background="#FDFEFE")
        self.col_tree.tag_configure("odd",     background="#EBF5FB")

        col_sb_v = ttk.Scrollbar(tree_fr, orient=tk.VERTICAL,   command=self.col_tree.yview)
        col_sb_h = ttk.Scrollbar(tree_fr, orient=tk.HORIZONTAL, command=self.col_tree.xview)
        self.col_tree.configure(yscrollcommand=col_sb_v.set, xscrollcommand=col_sb_h.set)
        self.col_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        col_sb_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        col_sb_h.grid(row=1, column=0, sticky=(tk.W, tk.E))

    def _populate_col_tree(self):
        """Rebuild the Column Lineage tree from self.targeted_results."""
        if not self.col_tree:
            return
        self._col_all_rows = []
        for result in self.targeted_results:
            tgt = result.get("target_table", "")
            for row in result.get("rows", []):
                if row.get("Record_Type") == "COLUMN_MAPPING":
                    self._col_all_rows.append((tgt, row))

        total    = len(self._col_all_rows)
        n_tables = len(self.targeted_results)
        if self.col_summary_var:
            self.col_summary_var.set(
                f"{n_tables} target table{'s' if n_tables != 1 else ''}  |  "
                f"{total} column mapping row{'s' if total != 1 else ''}"
            )
        if self.col_export_btn:
            self.col_export_btn.config(state=tk.NORMAL if total else tk.DISABLED)
        if self.col_filter_var:
            self.col_filter_var.set("")
        self._render_col_tree(self._col_all_rows)

    def _render_col_tree(self, rows):
        """Render (tgt, row) pairs into the column tree grouped by target table."""
        if not self.col_tree:
            return
        for item in self.col_tree.get_children():
            self.col_tree.delete(item)

        from collections import defaultdict
        groups: dict = defaultdict(list)
        for tgt, row in rows:
            groups[tgt].append(row)

        for tgt, group_rows in groups.items():
            parent = self.col_tree.insert(
                "", tk.END,
                text=f"{tgt}   ({len(group_rows)} mappings)",
                values=("", "", "", "", ""),
                tags=("target_hdr",),
                open=True,
            )
            for idx, row in enumerate(group_rows):
                src_col  = row.get("Source_Column", "")
                tgt_col  = row.get("Target_Column", "")
                src_tbl  = row.get("Source_Object", "")
                expr_raw = (row.get("Formula_Clean") or row.get("Formula") or "").strip()
                expr_disp = (expr_raw[:58] + "…") if len(expr_raw) > 58 else expr_raw
                df_name  = row.get("Dataflow_Name", "")
                job      = row.get("Job_Name", "")
                label    = (f"{src_col}  →  {tgt_col}"
                            if src_col and tgt_col else src_col or tgt_col or "(no column)")
                self.col_tree.insert(
                    parent, tk.END,
                    text=label,
                    values=(src_tbl, expr_disp, tgt_col, df_name, job),
                    tags=("odd" if idx % 2 else "mapping",),
                )

    def _apply_col_filter(self):
        """Re-render column tree filtered by the current filter text."""
        if not self._col_all_rows:
            return
        ftext = (self.col_filter_var.get() if self.col_filter_var else "").strip().upper()
        if not ftext:
            self._render_col_tree(self._col_all_rows)
            return
        filtered = [
            (tgt, row) for tgt, row in self._col_all_rows
            if (ftext in tgt.upper()
                or ftext in row.get("Source_Column", "").upper()
                or ftext in row.get("Target_Column", "").upper()
                or ftext in row.get("Source_Object", "").upper()
                or ftext in row.get("Dataflow_Name", "").upper())
        ]
        self._render_col_tree(filtered)

    def _export_col_csv(self):
        """Export all column lineage rows to a CSV file."""
        if not self._col_all_rows:
            messagebox.showinfo("Nothing to Export", "Run a Targeted Search first.")
            return
        import csv
        out_dir  = self.targeted_output_var.get().strip() or os.path.expanduser("~/Documents")
        safe_nm  = re.sub(r"[^A-Za-z0-9_.-]", "_",
                          self.targeted_results[0]["target_table"]) if self.targeted_results else "lineage"
        def_path = os.path.join(out_dir, f"col_lineage_{safe_nm}.csv")
        path = filedialog.asksaveasfilename(
            title="Save Column Lineage CSV",
            initialdir=os.path.dirname(def_path),
            initialfile=os.path.basename(def_path),
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            defaultextension=".csv",
        )
        if not path:
            return
        fieldnames = [
            "Target_Table", "Source_Column", "Source_Object",
            "Formula_Clean", "Formula", "Target_Column",
            "Dataflow_Name", "Workflow_Name", "Job_Name",
        ]
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for tgt, row in self._col_all_rows:
                merged = {"Target_Table": tgt}
                merged.update(row)
                writer.writerow(merged)
        messagebox.showinfo("Export Complete", f"Column lineage saved to:\n{path}")

    # ── Job: helpers ─────────────────────────────────────────────────────

    def _browse_job_xml(self):
        current = self.job_xml_var.get()
        init_dir = os.path.dirname(current) if os.path.isfile(current) else (
            current if os.path.isdir(current) else os.getcwd()
        )
        path = filedialog.askopenfilename(
            title="Select BODS Repository XML",
            initialdir=init_dir,
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
        )
        if path:
            self.job_xml_var.set(path)

    def _browse_job_output(self):
        current = self.job_output_var.get()
        dir_path = filedialog.askdirectory(
            title="Select Output Directory",
            initialdir=current if os.path.isdir(current) else os.path.expanduser("~/Documents"),
        )
        if dir_path:
            self.job_output_var.set(dir_path)

    def _copy_main_path_to_job(self):
        """Copy the Full Extraction input path into the Job XML field."""
        p = self.input_path.get().strip()
        if p:
            self.job_xml_var.set(p)

    def _j_log_msg(self, msg: str, level: str = "info"):
        prefix = {"status": "[STATUS]", "info": "[INFO]", "warning": "[WARN]",
                  "error": "[ERROR]", "success": "[OK]"}.get(level, "[INFO]")
        self.j_log.after(0, self._j_log_append, f"{prefix} {msg}\n")

    def _j_log_append(self, text: str):
        self.j_log.insert(tk.END, text)
        self.j_log.see(tk.END)

    # ── Job: run ─────────────────────────────────────────────────────────

    def run_job_search(self):
        """Validate inputs and start the job lineage worker thread."""
        if self.job_is_running:
            return

        xml_path  = self.job_xml_var.get().strip()
        raw_input = self.job_name_var.get().strip()
        max_hops  = self.job_hops_var.get()

        if not xml_path or not os.path.isfile(xml_path):
            messagebox.showerror(
                "Input Folder Not Set",
                f"BODS_FULL_PROD.xml was not found.\n\n"
                f"Browse to the input folder (above the tabs) that contains:\n"
                f"  • {self._MAIN_XML}\n"
                f"  • {self._RAW_DF}  (optional)\n"
                f"  • {self._RAW_ABAP}  (optional)",
                parent=self.root,
            )
            return
        if not raw_input:
            messagebox.showerror("Error", "Enter one or more job names (comma-separated).", parent=self.root)
            return

        job_names = [j.strip() for j in raw_input.split(",") if j.strip()]

        self.j_log.delete(1.0, tk.END)
        for item in self.j_tree.get_children():
            self.j_tree.delete(item)
        n = len(job_names)
        self.j_summary_var.set(f"Running analysis for {n} job{'s' if n > 1 else ''}…")
        self.j_progress_var.set(0)
        self.j_progress_lbl.config(text="")

        self.job_results = []
        self.job_is_running = True
        self._job_start_time = time.time()
        self.j_run_btn.config(state=tk.DISABLED)
        self.j_stop_btn.config(state=tk.NORMAL)
        if self.job_graph_btn:
            self.job_graph_btn.config(state=tk.DISABLED)

        t = threading.Thread(
            target=self._run_job_thread,
            args=(xml_path, job_names, max_hops),
            daemon=True,
        )
        t.start()
        self.root.after(100, self._check_job_status)

    def _run_job_thread(self, xml_path: str, job_names: list, max_hops: int):
        """Worker thread — runs JobLineageRunner for each job name."""
        try:
            from ds_engine.job_lineage_runner import JobLineageRunner

            def cb(msg, level="info"):
                self.job_queue.put(("log", (msg, level)))
                if level == "status":
                    self.job_queue.put(("status", msg))

            # Build RawDumpStore from all raw dump files found in the input folder
            raw_dump = None
            rd_paths = [p for p in self.resolved_raw_dump_paths if os.path.isfile(p)]
            if rd_paths:
                try:
                    from ds_engine.ds_raw_dump_index import RawDumpStore
                    raw_dump = RawDumpStore(rd_paths, status_callback=cb)
                    raw_dump.build_index()
                except Exception as exc:
                    self.job_queue.put((
                        "log",
                        (f"[RawDump] Failed to index raw dumps: {exc}", "warning"),
                    ))
                    raw_dump = None

            n = len(job_names)
            for idx, job_name in enumerate(job_names):
                if not self.job_is_running:
                    break
                pct_start = 5.0 + (idx / n) * 90.0
                self.job_queue.put(("progress", pct_start))
                self.job_queue.put(("log", (f"[{idx+1}/{n}] Analysing job: {job_name}", "status")))

                runner = JobLineageRunner(xml_path, status_callback=cb)
                results = runner.run(job_name, max_hops=max_hops, raw_dump=raw_dump)

                for result in results:
                    result["_job_name"] = job_name
                    self.job_queue.put(("result", result))

                self.job_queue.put(("progress", 5.0 + ((idx + 1) / n) * 90.0))

        except Exception as exc:
            import traceback
            self.job_queue.put(("log", (f"FATAL: {exc}", "error")))
            for line in traceback.format_exc().splitlines():
                if line.strip():
                    self.job_queue.put(("log", (line, "error")))
        finally:
            self.job_queue.put(("done", None))

    def _check_job_status(self):
        """Poll the job queue and update UI — called via root.after."""
        try:
            while True:
                msg_type, payload = self.job_queue.get_nowait()
                if msg_type == "log":
                    msg, level = payload
                    self._j_log_msg(msg, level)
                elif msg_type == "status":
                    self.j_summary_var.set(payload)
                elif msg_type == "progress":
                    pct = float(payload)
                    self.j_progress_var.set(pct)
                    self.j_progress_lbl.config(text=f"{pct:.1f}%")
                elif msg_type == "result":
                    self.job_results.append(payload)
                elif msg_type == "done":
                    self._job_finished()
                    return
        except queue.Empty:
            pass

        if self.job_is_running:
            self.root.after(100, self._check_job_status)

    def _job_finished(self):
        """Called when the job worker thread signals done."""
        self.job_is_running = False
        self.j_run_btn.config(state=tk.NORMAL)
        self.j_stop_btn.config(state=tk.DISABLED)
        self.j_progress_var.set(100)
        self.j_progress_lbl.config(text="100.0%")

        elapsed = time.time() - self._job_start_time
        m, s = divmod(int(elapsed), 60)
        elapsed_str = f"{m}m {s}s" if m else f"{s}s"

        if self.job_results:
            n           = len(self.job_results)
            total_terms = sum(r.get("stats", {}).get("terminal_srcs", 0) for r in self.job_results)
            total_rows  = sum(r.get("stats", {}).get("total_rows",   0) for r in self.job_results)
            tables_str  = ", ".join(r["target_table"] for r in self.job_results[:5])
            if n > 5:
                tables_str += f"… (+{n - 5} more)"
            self.j_summary_var.set(
                f"{n} target table{'s' if n > 1 else ''}  |  "
                f"{total_terms} terminal source(s)  |  {total_rows} column row(s)  |  {elapsed_str}"
            )
            for item in self.j_tree.get_children():
                self.j_tree.delete(item)
            for result in self.job_results:
                self._populate_job_tree(result)
            if self.job_graph_btn:
                self.job_graph_btn.config(state=tk.NORMAL)
            # Prepend to recent-searches combobox (deduplicated, capped at 15)
            search_str = self.job_name_var.get().strip()
            if search_str and (not self.job_recent_tables or self.job_recent_tables[0] != search_str):
                self.job_recent_tables = (
                    [search_str] + [x for x in self.job_recent_tables if x != search_str]
                )[:15]
                self.j_name_entry['values'] = self.job_recent_tables
        else:
            self.j_summary_var.set(
                f"Job analysis finished — no targets found (check log for details).  |  {elapsed_str}"
            )
        self._save_session_state()

    def _stop_job(self):
        if self.job_is_running:
            self.job_is_running = False
            self.job_queue.put(("log", ("Analysis stopped by user.", "warning")))
            self._job_finished()

    def _clear_job(self):
        self.j_log.delete(1.0, tk.END)
        for item in self.j_tree.get_children():
            self.j_tree.delete(item)
        self.j_summary_var.set("No job analysis run yet.")
        self.j_progress_var.set(0)
        self.j_progress_lbl.config(text="")
        self.job_results = []

    # ── Job: tree population ─────────────────────────────────────────────

    def _populate_job_tree(self, result: dict):
        """Append one target's lineage subtree to the job tree (no clear)."""
        if not result:
            return

        target_upper  = result["target_table"].upper()
        upstream_tree = result.get("upstream_tree", {})
        hop_map       = result.get("hop_map", {})
        terminals     = {t.upper() for t in result.get("terminal_sources", set())}
        table_context = result.get("table_context", {})

        # Insert the target table as a root node in the tree
        ctx = table_context.get(target_upper, {})
        root_id = self.j_tree.insert(
            "", "end",
            text=result["target_table"],
            values=(
                "0", "TARGET",
                ctx.get("Job_Name", ""),
                ctx.get("Workflow_Name", ""),
                ctx.get("Dataflow_Name", ""),
            ),
            open=True,
            tags=("target",),
        )

        def add_node(parent_id, table: str, path: frozenset):
            tup = table.upper()
            if tup in path:
                return
            if tup == target_upper:
                # already inserted as root
                for child in upstream_tree.get(tup, []):
                    add_node(root_id, child, path | {tup})
                return
            hop = hop_map.get(tup, "?")
            if tup in terminals:
                node_type, tag = "FINAL SOURCE", "source"
            else:
                node_type, tag = "INTERMEDIATE", "intermediate"

            node_ctx = table_context.get(tup, {})
            item_id = self.j_tree.insert(
                parent_id, "end",
                text=table,
                values=(
                    str(hop), node_type,
                    node_ctx.get("Job_Name", ""),
                    node_ctx.get("Workflow_Name", ""),
                    node_ctx.get("Dataflow_Name", ""),
                ),
                open=(isinstance(hop, int) and hop < 2),
                tags=(tag,),
            )
            for child in upstream_tree.get(tup, []):
                add_node(item_id, child, path | {tup})

        add_node(root_id, target_upper, frozenset())

    # ── Job: graph ───────────────────────────────────────────────────────

    def _open_job_graph(self):
        """Build the multi-tab HTML graph for all job targets and open in browser."""
        if not self.job_results:
            return
        import webbrowser
        try:
            path = self._build_job_lineage_html(self.job_results)
            webbrowser.open("file:///" + path.replace(os.sep, "/"))
        except Exception as exc:
            messagebox.showerror("Graph Error", str(exc), parent=self.root)

    @staticmethod
    @staticmethod
    def _group_results_by_dataflow(results: list) -> list:
        """One dataset entry per target-table result — each dataflow gets its own tab.

        Tab label  = Dataflow_Name (from table_context) so the tab strip reads as a
                     clean list of dataflow names rather than one merged job blob.
        Tab tooltip = full target-table name.
        """
        datasets = []
        for r in results:
            recs = DSXMLLauncher._result_to_records(r)
            if not recs:
                continue
            col_data = DSXMLLauncher._result_to_col_data(r)
            tgt = r["target_table"]
            # Look up this target's dataflow name from table_context
            ctx = r.get("table_context", {})
            tgt_ctx = ctx.get(tgt.upper()) or ctx.get(tgt) or {}
            df_name = tgt_ctx.get("Dataflow_Name") or ""
            # Build sql_queries the same way _build_lineage_html does
            sq_raw = r.get("sql_queries", {})
            sql_queries = {
                **{
                    f"SQL:{(row.get('Transform_Name') or row.get('Dataflow_Name') or 'SQL').strip()}": (row.get("Formula") or "").strip()
                    for row in r.get("rows", [])
                    if row.get("Record_Type") == "SQL_TRANSFORM"
                    and (row.get("Formula") or "").strip()
                },
                **sq_raw,
                **{
                    f"SQL:{ck[4:].split('|')[-1]}": qt
                    for ck, qt in sq_raw.items()
                    if "|" in ck[4:]
                },
            }
            datasets.append({
                "tgt": tgt,
                "label": df_name or tgt.split(".")[-1],
                "recs": recs,
                "cols": col_data,
                "sql_members": r.get("sql_members", {}),
                "sql_queries": sql_queries,
            })
        return datasets

    @staticmethod
    def _group_results_by_job(results: list) -> list:
        """Legacy: merge per-table results into one entry per job.

        Kept for reference; _build_job_lineage_html now calls
        _group_results_by_dataflow instead so each dataflow gets its own tab.
        """
        from collections import OrderedDict

        grouped: "OrderedDict[str, list]" = OrderedDict()
        for r in results:
            key = r.get("_job_name") or r.get("target_table", "unknown")
            grouped.setdefault(key, []).append(r)

        datasets = []
        for job_name, job_results in grouped.items():
            all_recs: list = []
            merged_cols: dict = {}
            targets: list = []
            for r in job_results:
                recs = DSXMLLauncher._result_to_records(r)
                if recs:
                    all_recs.extend(recs)
                col_data = DSXMLLauncher._result_to_col_data(r)
                merged_cols.update(col_data)
                targets.append(r["target_table"])
            if all_recs:
                datasets.append({
                    "tgt": job_name,
                    "label": job_name,
                    "targets": targets,
                    "recs": all_recs,
                    "cols": merged_cols,
                })
        return datasets

    def _build_job_lineage_html(self, results: list) -> str:
        """Generate a self-contained multi-tab HTML lineage graph for job results.

        Each dataflow in the job gets its own tab (tab label = dataflow name).
        """
        import json

        datasets = self._group_results_by_dataflow(results)

        if not datasets:
            raise ValueError(
                "No lineage edges found for any target in this job. "
                "Run the job analysis first."
            )

        html = _LINEAGE_HTML_TEMPLATE.replace("%%DATASETS%%", json.dumps(datasets))

        import time as _time
        job_label = re.sub(r"[^A-Za-z0-9_.-]", "_", self.job_name_var.get().strip().split(",")[0])
        fname = f"job_lineage_{job_label[:40]}_{int(_time.time())}.html"
        out_dir = self.job_output_var.get().strip()
        if not out_dir or not os.path.isdir(out_dir):
            xml_path = self.job_xml_var.get().strip()
            if xml_path and os.path.isfile(xml_path):
                out_dir = os.path.dirname(os.path.abspath(xml_path))
            else:
                out_dir = os.path.expanduser(os.path.join("~", "Documents"))
                os.makedirs(out_dir, exist_ok=True)

        out_path = os.path.join(out_dir, fname)
        with open(out_path, "w", encoding="utf-8") as fh:
            fh.write(html)
        return out_path

    # ── Job: export ──────────────────────────────────────────────────────

    def _export_job_column_csv(self):
        if not self.job_results:
            messagebox.showwarning("No Data", "Run a job analysis first.", parent=self.root)
            return
        all_rows = [r for res in self.job_results for r in res.get("rows", [])]
        if not all_rows:
            messagebox.showwarning("No Data", "No column-level rows to export.", parent=self.root)
            return
        job_label = re.sub(r"[^A-Za-z0-9_.-]", "_", self.job_name_var.get().strip().split(",")[0])
        out_dir = self.job_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, f"job_column_{job_label}.csv")
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Column Lineage CSV",
                defaultextension=".csv",
                initialfile=f"job_column_{job_label}.csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            pd.DataFrame(all_rows).to_csv(fp, index=False)
            messagebox.showinfo("Exported", f"Column lineage saved:\n{fp}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)

    def _export_job_table_csv(self):
        if not self.job_results:
            messagebox.showwarning("No Data", "Run a job analysis first.", parent=self.root)
            return
        job_label = re.sub(r"[^A-Za-z0-9_.-]", "_", self.job_name_var.get().strip().split(",")[0])
        out_dir = self.job_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, f"job_table_{job_label}.csv")
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Table Lineage CSV",
                defaultextension=".csv",
                initialfile=f"job_table_{job_label}.csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            records = []
            for result in self.job_results:
                hop_map       = result.get("hop_map", {})
                upstream_tree = result.get("upstream_tree", {})
                terminals     = {t.upper() for t in result.get("terminal_sources", set())}
                table_context = result.get("table_context", {})
                for tgt, srcs in upstream_tree.items():
                    ctx     = table_context.get(tgt.upper(), {})
                    tgt_hop = hop_map.get(tgt.upper(), "")
                    for src in srcs:
                        records.append({
                            "Target_Table":  tgt,
                            "Source_Table":  src,
                            "Target_Hop":    tgt_hop,
                            "Source_Hop":    hop_map.get(src.upper(), ""),
                            "Job_Name":      ctx.get("Job_Name", ""),
                            "Workflow_Name": ctx.get("Workflow_Name", ""),
                            "Dataflow_Name": ctx.get("Dataflow_Name", ""),
                        })
            if not records:
                messagebox.showwarning("No Data", "No table lineage data to export.", parent=self.root)
                return
            pd.DataFrame(records).to_csv(fp, index=False)
            messagebox.showinfo("Exported", f"Table lineage saved:\n{fp}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)

    def _export_job_excel(self):
        if not self.job_results:
            messagebox.showwarning("No Data", "Run a job analysis first.", parent=self.root)
            return
        job_label    = re.sub(r"[^A-Za-z0-9_.-]", "_", self.job_name_var.get().strip().split(",")[0])
        default_name = f"job_lineage_{job_label}.xlsx"
        out_dir = self.job_output_var.get().strip()
        if out_dir and os.path.isdir(out_dir):
            fp = os.path.join(out_dir, default_name)
        else:
            fp = filedialog.asksaveasfilename(
                title="Save Excel Workbook",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
        if not fp:
            return
        try:
            from datetime import datetime as _dt

            with pd.ExcelWriter(fp, engine="openpyxl") as writer:
                summary_rows = []
                for result in self.job_results:
                    stats = result.get("stats", {})
                    summary_rows.append({
                        "Target_Table":      result["target_table"],
                        "Total_Column_Rows": stats.get("total_rows", ""),
                        "Hops":              stats.get("hops", ""),
                        "Dataflows":         stats.get("dataflows", ""),
                        "Terminal_Sources":  stats.get("terminal_srcs", ""),
                        "Jobs":              stats.get("jobs", ""),
                        "Elapsed_s":         stats.get("elapsed_s", ""),
                        "Generated":         _dt.now().strftime("%Y-%m-%d %H:%M"),
                    })
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

                for result in self.job_results:
                    hop_map       = result.get("hop_map", {})
                    upstream_tree = result.get("upstream_tree", {})
                    terminals     = {t.upper() for t in result.get("terminal_sources", set())}
                    table_context = result.get("table_context", {})
                    rows          = result.get("rows", [])

                    table_records = []
                    for tgt, srcs in upstream_tree.items():
                        ctx      = table_context.get(tgt.upper(), table_context.get(tgt, {}))
                        tgt_hop  = hop_map.get(tgt.upper(), hop_map.get(tgt, ""))
                        tgt_type = "TARGET" if tgt_hop == 0 else "INTERMEDIATE"
                        for src in srcs:
                            src_hop  = hop_map.get(src.upper(), hop_map.get(src, ""))
                            src_type = "FINAL SOURCE" if src.upper() in terminals else "INTERMEDIATE"
                            table_records.append({
                                "Lineage_Path":  f"{src}  -->  {tgt}  (via {ctx.get('Dataflow_Name','?').split(';')[0].strip()} | {ctx.get('Job_Name','?').split(';')[0].strip()})",
                                "Target_Table":  tgt, "Target_Type": tgt_type, "Target_Hop": tgt_hop,
                                "Source_Table":  src, "Source_Type": src_type, "Source_Hop": src_hop,
                                "Job_Name":      ctx.get("Job_Name", ""),
                                "Workflow_Name": ctx.get("Workflow_Name", ""),
                                "Dataflow_Name": ctx.get("Dataflow_Name", ""),
                            })

                    safe_sheet = re.sub(r"[^A-Za-z0-9_]", "_", result["target_table"])[:28]
                    col_df = pd.DataFrame(rows) if rows else pd.DataFrame()
                    if not col_df.empty:
                        col_df.to_excel(writer, sheet_name=f"{safe_sheet}_Col", index=False)
                    if table_records:
                        tdf = pd.DataFrame(table_records)
                        _smap = {"TARGET": 0, "INTERMEDIATE": 1, "FINAL SOURCE": 2}
                        tdf["_s"] = tdf["Source_Type"].map(_smap).fillna(1)
                        tdf.sort_values(["Target_Hop", "Source_Hop", "_s"], inplace=True)
                        tdf.drop(columns=["_s"], inplace=True)
                        tdf.reset_index(drop=True, inplace=True)
                        tdf.to_excel(writer, sheet_name=f"{safe_sheet}_Tbl", index=False)
                        self._format_table_lineage_sheet(
                            writer.sheets[f"{safe_sheet}_Tbl"], tdf)

            messagebox.showinfo("Exported", f"Excel workbook saved:\n{fp}", parent=self.root)
        except ImportError:
            messagebox.showerror("Export Error", "openpyxl not installed. Run: pip install openpyxl", parent=self.root)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self.root)


def main():
    """Main entry point"""
    root = tk.Tk()
    app = DSXMLLauncher(root)
    root.mainloop()


if __name__ == "__main__":
    main()